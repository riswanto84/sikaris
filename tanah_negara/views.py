from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.db.models.deletion import ProtectedError
from django.views.generic import CreateView, UpdateView, DetailView, DeleteView, FormView

from core.forms import ImportFileForm
from core.import_utils import read_tabular_upload, pick, to_decimal, to_date, normalize_choice
from core.export_utils import apply_search_filter, export_queryset
from core.listing import SearchListMixin
from core.roles import AdminSystemRequiredMixin
from master.models import UnitKerja
from .forms import TanahNegaraForm
from .models import TanahNegara


class SafeDeleteMixin:
    template_name = 'includes/confirm_delete.html'
    success_message = 'Data berhasil dihapus.'
    protected_message = 'Data tidak dapat dihapus karena masih digunakan oleh data lain.'

    def form_valid(self, form):
        try:
            messages.success(self.request, self.success_message)
            return super().form_valid(form)
        except ProtectedError:
            messages.error(self.request, self.protected_message)
            return redirect(self.get_success_url())


class TanahNegaraListView(AdminSystemRequiredMixin, SearchListMixin):
    model = TanahNegara
    template_name = 'tanah_negara/list.html'
    select_related = ['unit_kerja']
    search_fields = [
        ('kode_tanah','Kode Tanah'), ('kode_satker','Kode Satker'), ('nama_satker','Nama Satker'),
        ('nama_tanah','Nama Tanah'), ('nama_aset','Nama Aset'), ('nama_barang','Nama Barang'),
        ('unit_kerja__nama_unit','Unit Kerja'), ('kode_barang','Kode Barang'), ('nup','NUP'),
        ('kode_register','Kode Register'), ('alamat','Alamat'), ('rt_rw','RT/RW'),
        ('kelurahan_desa','Kelurahan/Desa'), ('kecamatan','Kecamatan'), ('kab_kota','Kab/Kota'),
        ('provinsi','Provinsi'), ('nomor_sertifikat','Nomor Sertifikat'), ('nomor_dokumen','Nomor Dokumen'),
        ('jenis_dokumen','Jenis Dokumen'), ('status_sertifikasi','Status Sertifikasi'), ('status_tanah','Kategori Status'),
        ('status_bmn','Status BMN'), ('status_penggunaan','Status Penggunaan'), ('status_pemanfaatan','Status Pemanfaatan'),
        ('penghuni','Penghuni'), ('pengguna','Pengguna'), ('digunakan_oleh','Digunakan Oleh'), ('keterangan','Keterangan')
    ]


class TanahNegaraCreateView(AdminSystemRequiredMixin, CreateView):
    model = TanahNegara
    form_class = TanahNegaraForm
    template_name = 'tanah_negara/form.html'
    success_url = reverse_lazy('tanah_negara:list')

    def form_valid(self, form):
        form.instance.dibuat_oleh = self.request.user
        if not form.instance.nama_tanah and form.instance.nama_aset:
            form.instance.nama_tanah = form.instance.nama_aset
        messages.success(self.request, 'Data tanah negara berhasil disimpan.')
        return super().form_valid(form)


class TanahNegaraUpdateView(AdminSystemRequiredMixin, UpdateView):
    model = TanahNegara
    form_class = TanahNegaraForm
    template_name = 'tanah_negara/form.html'
    success_url = reverse_lazy('tanah_negara:list')

    def form_valid(self, form):
        if not form.instance.nama_tanah and form.instance.nama_aset:
            form.instance.nama_tanah = form.instance.nama_aset
        messages.success(self.request, 'Data tanah negara berhasil diperbarui.')
        return super().form_valid(form)


class TanahNegaraDetailView(AdminSystemRequiredMixin, DetailView):
    model = TanahNegara
    template_name = 'tanah_negara/detail.html'


class TanahNegaraDeleteView(AdminSystemRequiredMixin, SafeDeleteMixin, DeleteView):
    model = TanahNegara
    success_url = reverse_lazy('tanah_negara:list')
    success_message = 'Tanah negara berhasil dihapus.'


def _unit(name):
    if not name:
        return None
    return UnitKerja.objects.get_or_create(nama_unit=str(name).strip())[0]


def _status_tanah(row):
    text = ' '.join(str(pick(row, key, default='')) for key in [
        'status_tanah', 'status_pemanfaatan', 'status_penggunaan', 'status_bmn', 'status_pmk', 'keterangan'
    ]).lower()
    if 'sengketa' in text or 'gugatan' in text or 'perkara' in text:
        return 'SENGKETA'
    if 'idle' in text or 'kosong' in text:
        return 'IDLE'
    if 'sewa' in text or 'dimanfaatkan' in text or 'pihak lain' in text:
        return 'DISEWAKAN'
    if 'digunakan' in text or 'aktif' in text or 'operasional' in text:
        return 'DIGUNAKAN'
    return normalize_choice(pick(row, 'status_tanah'), TanahNegara.STATUS_TANAH, 'DIGUNAKAN')




# =============================================================
# Template Excel Import Tanah Negara
# =============================================================
def template_import_tanah(request):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        return HttpResponse(f'openpyxl belum tersedia: {exc}', status=500, content_type='text/plain')

    headers = [
        'kode_tanah', 'kode_satker', 'nama_satker', 'unit_kerja', 'kode_barang', 'nup',
        'nama_barang', 'nama_aset', 'nama_tanah', 'status_bmn', 'kondisi', 'intra_extra',
        'jenis_dokumen', 'nomor_dokumen', 'status_sertifikasi', 'jenis_sertipikat',
        'nomor_sertifikat', 'atas_nama_sertifikat', 'tanggal_sertifikat', 'tanggal_buku_pertama',
        'tanggal_perolehan', 'nilai_perolehan', 'nilai_buku', 'luas_tanah',
        'luas_tanah_untuk_bangunan', 'luas_tanah_untuk_sarana_lingkungan', 'luas_lahan_kosong',
        'luas_pemanfaatan', 'status_penggunaan', 'status_pemanfaatan', 'status_tanah',
        'no_psp', 'tanggal_psp', 'alamat', 'rt_rw', 'kelurahan_desa', 'kecamatan',
        'kabupaten_kota', 'provinsi', 'kode_pos', 'latitude', 'longitude', 'penghuni',
        'pengguna', 'digunakan_oleh', 'kode_kpknl', 'uraian_kpknl', 'kode_register',
        'status_pmk', 'keterangan'
    ]
    sample = [
        'TN-001', '027.01.01', 'Biro Umum', 'Biro Umum', '2.01.01.01.001', '001',
        'Tanah Bangunan Kantor Pemerintah', 'Tanah Kantor Contoh', 'Tanah Kantor Contoh',
        'Aktif', 'Baik', 'Intra', 'Sertifikat', 'SHM-001', 'Sertifikat', 'SHM',
        'SHM-001', 'Kementerian Sosial RI', '2020-01-01', '2020-01-01', '2020-01-01',
        1500000000, 1500000000, 1000, 600, 200, 200, 0, 'Digunakan', 'Digunakan',
        'Digunakan', 'PSP-001', '2021-01-01', 'Jl. Contoh No. 1', '001/002', 'Senen',
        'Senen', 'Jakarta Pusat', 'DKI Jakarta', '10410', -6.175392, 106.827153,
        '', 'Biro Umum', 'Biro Umum', 'KPKNL-JKT', 'KPKNL Jakarta', 'REG-001', 'Aktif',
        'Contoh data tanah negara'
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Template Import'
    ws['A1'] = 'Template Import Tanah Negara'
    ws['A1'].font = Font(bold=True, size=14, color='0F172A')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.append([])
    ws.append(headers)
    ws.append(sample)

    header_fill = PatternFill('solid', fgColor='1D4ED8')
    header_font = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    for cell in ws[4]:
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=True)
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = min(max(len(str(header)) + 4, 14), 34)
    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:{get_column_letter(len(headers))}4'

    info = wb.create_sheet('Petunjuk')
    info['A1'] = 'Petunjuk Import Tanah Negara'
    info['A1'].font = Font(bold=True, size=14)
    info['A3'] = 'Isi data mulai baris ke-4. Jangan mengubah header pada baris ke-3.'
    info['A4'] = 'Format tanggal: YYYY-MM-DD. Nilai rupiah dan luas diisi angka saja.'
    info['A5'] = 'Kolom latitude/longitude boleh dikosongkan jika belum tersedia.'
    info.column_dimensions['A'].width = 110

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="template_import_tanah_negara.xlsx"'
    wb.save(response)
    return response


class TanahNegaraImportView(AdminSystemRequiredMixin, FormView):
    form_class = ImportFileForm
    template_name = 'includes/import_form.html'
    success_url = reverse_lazy('tanah_negara:list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update({
            'title': 'Impor Tanah Negara',
            'description': 'Upload Excel/CSV sesuai template: kode_tanah, kode_satker, nama_satker, unit_kerja, kode_barang, nup, nama_barang, nama_aset, status_bmn, kondisi, data sertifikat, luas, PSP, alamat, KPKNL, kode_register, status_pmk, keterangan. Latitude/longitude tetap didukung jika tersedia.',
            'back_url': reverse_lazy('tanah_negara:list'),
            'template_url': reverse_lazy('tanah_negara:template_import'),
        })
        return ctx

    def form_invalid(self, form):
        message = 'Form impor tidak valid. Pastikan file Excel/CSV sudah dipilih.'
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'ok': False, 'message': message}, status=400)
        messages.error(self.request, message)
        return redirect(self.success_url)

    def form_valid(self, form):
        created = updated = errors = 0
        try:
            rows = read_tabular_upload(form.cleaned_data['file'])
            for row in rows:
                kode = pick(row, 'kode_tanah', 'kode_aset', 'kode', 'kode_register', 'nup')
                nama_aset = pick(row, 'nama_aset', 'nama_tanah', 'nama_barang', 'nama')
                nama_tanah = pick(row, 'nama_tanah', 'nama_aset', 'nama_barang', 'nama', default=str(kode or ''))
                if not kode or not nama_tanah:
                    errors += 1
                    continue

                luas_seluruh = to_decimal(pick(row, 'luas_tanah_seluruhnya', 'luas_tanah', 'luas'), 0)
                nomor_sertifikat = pick(row, 'nomor_sertifikat', 'no_sertifikat', 'nomor_dokumen', 'no_dokumen')
                pengguna = pick(row, 'pengguna')
                penghuni = pick(row, 'penghuni')

                obj, is_created = TanahNegara.objects.update_or_create(
                    kode_tanah=str(kode).strip(),
                    defaults={
                        'kode_satker': pick(row, 'kode_satker'),
                        'nama_satker': pick(row, 'nama_satker'),
                        'unit_kerja': _unit(pick(row, 'unit_kerja', 'nama_satker', 'satker')),
                        'kode_barang': pick(row, 'kode_barang'),
                        'nup': pick(row, 'nup'),
                        'nama_barang': pick(row, 'nama_barang'),
                        'nama_aset': nama_aset,
                        'nama_tanah': nama_tanah,
                        'status_bmn': pick(row, 'status_bmn'),
                        'kondisi': pick(row, 'kondisi'),
                        'intra_extra': pick(row, 'intra_extra', 'intra__extra', 'intra/extra'),
                        'jenis_dokumen': pick(row, 'jenis_dokumen'),
                        'nomor_dokumen': pick(row, 'nomor_dokumen', 'no_dokumen'),
                        'status_sertifikasi': pick(row, 'status_sertifikasi'),
                        'jenis_sertipikat': pick(row, 'jenis_sertipikat', 'jenis_sertifikat'),
                        'nomor_sertifikat': nomor_sertifikat,
                        'atas_nama_sertifikat': pick(row, 'atas_nama_sertifikat'),
                        'tanggal_sertifikat': to_date(pick(row, 'tanggal_sertifikat')),
                        'tanggal_buku_pertama': to_date(pick(row, 'tanggal_buku_pertama')),
                        'tanggal_perolehan': to_date(pick(row, 'tanggal_perolehan')),
                        'nilai_perolehan': to_decimal(pick(row, 'nilai_perolehan', 'nilai'), 0),
                        'nilai_buku': to_decimal(pick(row, 'nilai_buku'), 0),
                        'luas_tanah': luas_seluruh,
                        'luas_tanah_seluruhnya': luas_seluruh,
                        'luas_tanah_untuk_bangunan': to_decimal(pick(row, 'luas_tanah_untuk_bangunan'), 0),
                        'luas_tanah_untuk_sarana_lingkungan': to_decimal(pick(row, 'luas_tanah_untuk_sarana_lingkungan'), 0),
                        'luas_lahan_kosong': to_decimal(pick(row, 'luas_lahan_kosong'), 0),
                        'luas_pemanfaatan': to_decimal(pick(row, 'luas_pemanfaatan'), 0),
                        'jumlah_foto': int(to_decimal(pick(row, 'jumlah_foto'), 0) or 0),
                        'status_penggunaan': pick(row, 'status_penggunaan'),
                        'status_pemanfaatan': pick(row, 'status_pemanfaatan'),
                        'status_tanah': _status_tanah(row),
                        'no_psp': pick(row, 'no_psp'),
                        'tanggal_psp': to_date(pick(row, 'tanggal_psp')),
                        'alamat': pick(row, 'alamat'),
                        'rt_rw': pick(row, 'rt_rw', 'rt/rw'),
                        'kelurahan_desa': pick(row, 'kelurahan_desa', 'kelurahan/desa'),
                        'kelurahan': pick(row, 'kelurahan', 'kelurahan_desa', 'kelurahan/desa'),
                        'kecamatan': pick(row, 'kecamatan'),
                        'kab_kota': pick(row, 'kab_kota', 'kab/kota', 'kabupaten_kota', 'kabupaten', 'kota'),
                        'kabupaten_kota': pick(row, 'kabupaten_kota', 'kab_kota', 'kab/kota', 'kabupaten', 'kota'),
                        'kode_kab_kota': pick(row, 'kode_kab_kota', 'kode_kab/kota'),
                        'provinsi': pick(row, 'provinsi'),
                        'kode_provinsi': pick(row, 'kode_provinsi'),
                        'kode_pos': pick(row, 'kode_pos'),
                        'latitude': to_decimal(pick(row, 'latitude', 'lat'), None),
                        'longitude': to_decimal(pick(row, 'longitude', 'long', 'lng'), None),
                        'sbsk': to_decimal(pick(row, 'sbsk'), 0),
                        'optimalisasi': to_decimal(pick(row, 'optimalisasi'), 0),
                        'penghuni': penghuni,
                        'pengguna': pengguna,
                        'digunakan_oleh': pick(row, 'digunakan_oleh', 'pengguna', 'penghuni'),
                        'kode_kpknl': pick(row, 'kode_kpknl'),
                        'uraian_kpknl': pick(row, 'uraian_kpknl'),
                        'uraian_kanwil_djkn': pick(row, 'uraian_kanwil_djkn'),
                        'nama_kl': pick(row, 'nama_kl'),
                        'nama_e1': pick(row, 'nama_e1'),
                        'nama_korwil': pick(row, 'nama_korwil'),
                        'kode_register': pick(row, 'kode_register'),
                        'status_pmk': pick(row, 'status_pmk'),
                        'keterangan': pick(row, 'keterangan', 'catatan'),
                        'dibuat_oleh': self.request.user,
                    })
                created += 1 if is_created else 0
                updated += 0 if is_created else 1
        except Exception as exc:
            if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'ok': False, 'message': str(exc)}, status=400)
            messages.error(self.request, str(exc))
            return redirect(self.success_url)
        msg = f'Impor tanah negara selesai. Baru: {created}, Update: {updated}, Gagal: {errors}.'
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'ok': True, 'message': msg, 'redirect_url': str(self.success_url)})
        messages.success(self.request, msg)
        return redirect(self.success_url)


# =============================================================
# Export daftar Tanah Negara (PDF, Excel, CSV)
# =============================================================
def _tanah_negara_columns():
    return [
        ('No', '__no__'),
        ('Kode Tanah', 'kode_tanah'),
        ('Kode Satker', 'kode_satker'),
        ('Nama Satker', 'nama_satker'),
        ('Unit Kerja', 'unit_kerja__nama_unit'),
        ('Kode Barang', 'kode_barang'),
        ('NUP', 'nup'),
        ('Kode Register', 'kode_register'),
        ('Nama Tanah/Aset', 'nama_tanah'),
        ('Nama Barang', 'nama_barang'),
        ('Alamat', 'alamat'),
        ('Kelurahan/Desa', 'kelurahan_desa'),
        ('Kecamatan', 'kecamatan'),
        ('Kab/Kota', lambda o: getattr(o, 'kab_kota', None) or getattr(o, 'kabupaten_kota', '')),
        ('Provinsi', 'provinsi'),
        ('Luas Tanah', lambda o: getattr(o, 'luas_tanah_seluruhnya', None) or getattr(o, 'luas_tanah', '')),
        ('Nilai Perolehan', 'nilai_perolehan'),
        ('Status Tanah', 'display:status_tanah'),
        ('Status Penggunaan', 'status_penggunaan'),
        ('Status Pemanfaatan', 'status_pemanfaatan'),
        ('Nomor Sertifikat', 'nomor_sertifikat'),
        ('No PSP', 'no_psp'),
        ('Latitude', 'latitude'),
        ('Longitude', 'longitude'),
        ('Keterangan', 'keterangan'),
    ]


@login_required
def export_tanah_negara(request, fmt):
    if not (request.user.is_superuser or request.user.groups.filter(name='Admin System').exists()):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied('Anda tidak memiliki hak akses export Tanah Negara.')
    qs = TanahNegara.objects.select_related('unit_kerja')
    qs = apply_search_filter(qs, request, TanahNegaraListView.search_fields)
    return export_queryset(request, qs, fmt, 'master_tanah_negara', 'Master Tanah Negara', _tanah_negara_columns(), order_by=['kode_tanah'])

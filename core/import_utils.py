import csv
import io
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


def norm_header(value):
    return str(value or '').strip().lower().replace(' ', '_').replace('-', '_').replace('/', '_')


def read_tabular_upload(uploaded_file):
    name = (uploaded_file.name or '').lower()
    if name.endswith('.csv'):
        raw = uploaded_file.read().decode('utf-8-sig', errors='ignore')
        reader = csv.DictReader(io.StringIO(raw))
        return [{norm_header(k): (v or '').strip() for k, v in row.items()} for row in reader]
    if name.endswith(('.xlsx', '.xlsm')):
        if load_workbook is None:
            raise ValueError('openpyxl belum terinstall. Jalankan pip install openpyxl.')
        wb = load_workbook(uploaded_file, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [norm_header(h) for h in next(rows)]
        data = []
        for row in rows:
            if not any(v not in (None, '') for v in row):
                continue
            data.append({headers[i]: row[i] if i < len(row) else '' for i in range(len(headers))})
        return data
    raise ValueError('Format file harus .xlsx/.xlsm atau .csv')


def pick(row, *names, default=''):
    for name in names:
        key = norm_header(name)
        if key in row and row[key] not in (None, ''):
            return str(row[key]).strip() if not isinstance(row[key], (int, float, Decimal, date)) else row[key]
    return default


def to_int(value, default=None):
    if value in (None, ''):
        return default
    try:
        return int(float(str(value).replace(',', '.')))
    except Exception:
        return default


def to_decimal(value, default=0):
    if value in (None, ''):
        return default
    text = str(value).strip().replace('Rp', '').replace(' ', '')
    if text.count(',') == 1 and text.count('.') >= 1:
        text = text.replace('.', '').replace(',', '.')
    elif text.count(',') == 1:
        text = text.replace(',', '.')
    try:
        return Decimal(text)
    except InvalidOperation:
        return default


def to_date(value, default=None):
    if value in (None, ''):
        return default
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return default


def normalize_choice(value, choices, default=None):
    if value in (None, ''):
        return default
    text = str(value).strip().upper().replace(' ', '_').replace('-', '_')
    valid = {k: k for k, _ in choices}
    labels = {str(v).strip().upper().replace(' ', '_').replace('-', '_'): k for k, v in choices}
    if text in valid:
        return text
    return labels.get(text, default)

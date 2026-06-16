import csv
from xml.sax.saxutils import escape as xml_escape
from datetime import date, datetime
from decimal import Decimal

from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone


def stringify(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return timezone.localtime(value).strftime('%d/%m/%Y %H:%M') if timezone.is_aware(value) else value.strftime('%d/%m/%Y %H:%M')
    if isinstance(value, date):
        return value.strftime('%d/%m/%Y')
    if isinstance(value, Decimal):
        return f'{value:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    return str(value)


def get_nested_attr(obj, path):
    value = obj
    for part in path.split('__'):
        if value is None:
            return ''
        value = getattr(value, part, None)
        if callable(value) and not part.startswith('get_'):
            value = value()
    return value


def get_export_value(obj, accessor):
    if callable(accessor):
        return accessor(obj)
    # Special shortcut: display:status -> get_status_display()
    if isinstance(accessor, str) and accessor.startswith('display:'):
        field = accessor.split(':', 1)[1]
        method = getattr(obj, f'get_{field}_display', None)
        return method() if callable(method) else get_nested_attr(obj, field)
    return get_nested_attr(obj, accessor)


def apply_search_filter(qs, request, search_fields):
    q = (request.GET.get('q') or '').strip()
    selected_field = (request.GET.get('search_field') or 'ALL').strip()
    if q and search_fields:
        available_fields = [field for field, _label in search_fields]
        fields = available_fields if selected_field == 'ALL' or selected_field not in available_fields else [selected_field]
        query = Q()
        for field in fields:
            query |= Q(**{f'{field}__icontains': q})
        qs = qs.filter(query)
    return qs


def build_rows(qs, columns):
    rows = []
    for index, obj in enumerate(qs, start=1):
        row = []
        for header, accessor in columns:
            if accessor == '__no__':
                row.append(index)
            else:
                row.append(get_export_value(obj, accessor))
        rows.append(row)
    return rows


def export_csv(filename, headers, rows):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow([stringify(v) for v in row])
    return response


def export_excel(filename, sheet_title, headers, rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        return HttpResponse(f'openpyxl belum tersedia: {exc}', status=500, content_type='text/plain')

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    ws.append(headers)

    header_fill = PatternFill('solid', fgColor='1D4ED8')
    header_font = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    for row in rows:
        ws.append([stringify(v) for v in row])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = border

    for idx, header in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        max_len = len(str(header))
        for cell in ws[letter]:
            max_len = max(max_len, len(str(cell.value or '')))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 42)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_pdf(filename, title, headers, rows, landscape_mode=True):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except Exception as exc:
        return HttpResponse(f'reportlab belum tersedia: {exc}', status=500, content_type='text/plain')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    pagesize = landscape(A4) if landscape_mode else A4
    doc = SimpleDocTemplate(response, pagesize=pagesize, leftMargin=0.7*cm, rightMargin=0.7*cm, topMargin=0.8*cm, bottomMargin=0.8*cm)
    styles = getSampleStyleSheet()
    body = ParagraphStyle('ExportBody', parent=styles['BodyText'], fontSize=6.5, leading=8, wordWrap='CJK')
    header_style = ParagraphStyle('ExportHeader', parent=styles['BodyText'], fontSize=7, leading=8, textColor=colors.white, alignment=1, wordWrap='CJK')
    elements = [Paragraph(f'<b>{xml_escape(str(title))}</b>', styles['Title']), Paragraph(f'Dicetak: {timezone.localtime().strftime("%d/%m/%Y %H:%M")}', styles['Normal']), Spacer(1, 0.25*cm)]

    max_rows = 500
    table_rows = [[Paragraph(xml_escape(stringify(h)), header_style) for h in headers]]
    for row in rows[:max_rows]:
        table_rows.append([Paragraph(xml_escape(stringify(v)), body) for v in row])

    table = Table(table_rows, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1D4ED8')),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5E1')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(table)
    if len(rows) > max_rows:
        elements.append(Spacer(1, 0.2*cm))
        elements.append(Paragraph(f'Catatan: PDF dibatasi {max_rows} baris pertama. Gunakan Excel/CSV untuk data lengkap.', styles['Normal']))
    doc.build(elements)
    return response


def export_queryset(request, qs, fmt, filename_base, title, columns, order_by=None, landscape_mode=True):
    if order_by:
        qs = qs.order_by(*order_by) if isinstance(order_by, (list, tuple)) else qs.order_by(order_by)
    headers = [header for header, _accessor in columns]
    rows = build_rows(qs, columns)
    fmt = (fmt or 'xlsx').lower()
    if fmt == 'pdf':
        return export_pdf(f'{filename_base}.pdf', title, headers, rows, landscape_mode=landscape_mode)
    if fmt == 'csv':
        return export_csv(f'{filename_base}.csv', headers, rows)
    return export_excel(f'{filename_base}.xlsx', title, headers, rows)

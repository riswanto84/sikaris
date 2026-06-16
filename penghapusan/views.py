from decimal import Decimal, InvalidOperation
from io import BytesIO
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.db.models import Q
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import CreateView, UpdateView, DetailView, DeleteView
from openpyxl import Workbook, load_workbook

from core.access import is_biro_umum_user, is_global_bmn_scope_user, get_user_unit_kerja, require_user_unit_or_all, scope_queryset_by_user
from core.listing import SearchListMixin
from core.roles import is_sekretaris_jenderal, is_dirjen_rehsos
from .forms import ImportBarangPenghapusanForm, PermohonanPenghapusanBMNForm
from .models import BarangPenghapusanBMN, PermohonanPenghapusanBMN, FotoKondisiPenghapusanBMN


class PermohonanPenghapusanAccessMixin(LoginRequiredMixin):
    scope_type = 'penghapusan_bmn'

    def get_scoped_queryset(self):
        qs = PermohonanPenghapusanBMN.objects.select_related(
            'unit_kerja', 'pemohon', 'kendaraan', 'rumah_negara', 'tanah_negara',
            'dibuat_oleh', 'diverifikasi_oleh'
        )
        # Sekjen/Admin dapat melihat semua usulan penghapusan untuk penetapan SK.
        if is_sekretaris_jenderal(self.request.user) or self.request.user.is_superuser:
            return qs
        return scope_queryset_by_user(qs, self.request.user, 'penghapusan_bmn')

    def get_queryset(self):
        return self.get_scoped_queryset()

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs


class SafeDeleteMixin:
    template_name = 'includes/confirm_delete.html'
    success_message = 'Data berhasil dihapus.'
    protected_message = 'Data tidak dapat dihapus karena masih digunakan oleh data lain.'

    def form_valid(self, form):
        try:
            messages.success(self.request, self.success_message)
            return super().form_valid(form)
        except ProtectedError:
            messages.error(self.request, self.protected_message)
            return redirect(self.get_success_url())



class PermohonanPenghapusanListView(PermohonanPenghapusanAccessMixin, SearchListMixin):
    """Daftar usulan penghapusan dari unit kerja.

    Verifikasi Biro Umum dan penetapan SK Sekjen dipisahkan pada menu tersendiri.
    """
    model = PermohonanPenghapusanBMN
    template_name = 'penghapusan/list.html'
    select_related = ['unit_kerja', 'pemohon', 'kendaraan', 'rumah_negara', 'tanah_negara']
    mode = 'permohonan'
    page_title = 'Permohonan Penghapusan BMN'
    empty_message = 'Belum ada permohonan penghapusan BMN.'
    search_fields = [
        ('nomor_permohonan', 'Nomor Permohonan'),
        ('unit_kerja__nama_unit', 'Unit Kerja'),
        ('pemohon__nama', 'Nama Pemohon'),
        ('pemohon__nip', 'NIP Pemohon'),
        ('jenis_aset', 'Jenis Aset'),
        ('kode_barang', 'Kode Barang'),
        ('nup', 'NUP'),
        ('nama_barang', 'Nama Barang'),
        ('alasan_penghapusan', 'Alasan Penghapusan'),
        ('status', 'Status'),
        ('nomor_persetujuan', 'Nomor Persetujuan'),
        ('nomor_sk_penghapusan', 'Nomor SK Penghapusan'),
    ]

    def get_base_queryset_for_mode(self):
        if self.mode == 'verifikasi':
            if not (is_biro_umum_user(self.request.user) or self.request.user.is_superuser):
                raise PermissionDenied('Hanya role Biro Umum yang dapat mengakses Verifikasi Usulan Penghapusan.')
            return PermohonanPenghapusanBMN.objects.select_related(*self.select_related).exclude(status='DRAFT')
        if self.mode == 'persetujuan_dirjen_rehsos':
            if not (is_dirjen_rehsos(self.request.user) or self.request.user.is_superuser):
                raise PermissionDenied('Hanya role Dirjen Rehsos yang dapat mengakses Persetujuan Penghapusan Sentra Rehsos.')
            return PermohonanPenghapusanBMN.objects.select_related(*self.select_related).filter(status__in=['DIAJUKAN_KE_DIRJEN_REHSOS', 'DISETUJUI_DIRJEN_REHSOS', 'DITOLAK_DIRJEN_REHSOS'])
        if self.mode == 'persetujuan_sekjen':
            if not (is_sekretaris_jenderal(self.request.user) or self.request.user.is_superuser):
                raise PermissionDenied('Hanya role Sekjen yang dapat mengakses Penetapan Penghapusan BMN.')
            return PermohonanPenghapusanBMN.objects.select_related(*self.select_related).filter(status__in=['DIAJUKAN_KE_SEKJEN', 'DISETUJUI_DIRJEN_REHSOS', 'DITOLAK_SEKJEN', 'SK_PENGHAPUSAN_TERBIT', 'SELESAI', 'DIAJUKAN_SEKJEN', 'SK_TERBIT'])
        qs = self.get_scoped_queryset()
        if self.select_related:
            qs = qs.select_related(*self.select_related)
        return qs

    def get_queryset(self):
        qs = self.get_base_queryset_for_mode()
        q = (self.request.GET.get('q') or '').strip()
        selected_field = (self.request.GET.get('search_field') or 'ALL').strip()
        if q and self.search_fields:
            available_fields = [field for field, _label in self.search_fields]
            fields_to_search = available_fields if selected_field == 'ALL' or selected_field not in available_fields else [selected_field]
            query = Q()
            for field in fields_to_search:
                query |= Q(**{f'{field}__icontains': q})
            qs = qs.filter(query)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update({
            'mode': self.mode,
            'page_title': self.page_title,
            'empty_message': self.empty_message,
            'show_create_buttons': self.mode == 'permohonan' and not is_sekretaris_jenderal(self.request.user),
            'show_unit_edit_actions': self.mode == 'permohonan' and not is_sekretaris_jenderal(self.request.user),
            'show_dirjen_actions': self.mode == 'persetujuan_dirjen_rehsos',
        })
        return ctx


class VerifikasiPenghapusanListView(PermohonanPenghapusanListView):
    mode = 'verifikasi'
    page_title = 'Verifikasi Usulan Penghapusan BMN - Biro Umum'
    empty_message = 'Belum ada usulan penghapusan yang perlu diverifikasi Biro Umum.'


class PersetujuanSekjenPenghapusanListView(PermohonanPenghapusanListView):
    mode = 'persetujuan_sekjen'
    page_title = 'Penetapan SK Penghapusan BMN - Sekjen'
    empty_message = 'Belum ada usulan penghapusan yang diajukan ke Sekjen.'


class PersetujuanDirjenRehsosPenghapusanListView(PermohonanPenghapusanListView):
    mode = 'persetujuan_dirjen_rehsos'
    page_title = 'Persetujuan Penghapusan BMN - Dirjen Rehsos'
    empty_message = 'Belum ada usulan penghapusan Sentra Rehsos yang diajukan ke Dirjen Rehsos.'


def _save_foto_kondisi_files(request, permohonan):
    for foto in request.FILES.getlist('foto_kondisi_files'):
        FotoKondisiPenghapusanBMN.objects.create(
            permohonan=permohonan,
            foto=foto,
            diupload_oleh=request.user if request.user.is_authenticated else None,
        )

class PermohonanPenghapusanCreateView(PermohonanPenghapusanAccessMixin, CreateView):
    model = PermohonanPenghapusanBMN
    form_class = PermohonanPenghapusanBMNForm
    template_name = 'penghapusan/form.html'
    success_url = reverse_lazy('penghapusan:list')

    def form_valid(self, form):
        user = self.request.user
        if not is_global_bmn_scope_user(user):
            unit = get_user_unit_kerja(user)
            if not unit:
                raise PermissionDenied('User belum memiliki Unit Kerja/Satker.')
            form.instance.unit_kerja = unit
            form.instance.status = 'MENUNGGU_VERIFIKASI_BIRO_UMUM'
        else:
            if not form.instance.status:
                form.instance.status = 'MENUNGGU_VERIFIKASI_BIRO_UMUM'

        form.instance.dibuat_oleh = user
        form.instance.diperbarui_oleh = user
        self._fill_asset_snapshot(form.instance)
        response = super().form_valid(form)
        _save_foto_kondisi_files(self.request, self.object)
        messages.success(self.request, 'Permohonan penghapusan BMN berhasil diajukan.')
        return response

    def _fill_asset_snapshot(self, obj):
        if obj.jenis_aset == 'KENDARAAN' and obj.kendaraan:
            k = obj.kendaraan
            obj.kode_barang = obj.kode_barang or k.kode_barang
            obj.nup = obj.nup or k.nup
            obj.nama_barang = obj.nama_barang or f'{k.nomor_polisi} - {k.merek} {k.tipe or ""}'.strip()
            obj.nilai_perolehan = obj.nilai_perolehan or k.nilai_perolehan
            obj.kondisi_barang = obj.kondisi_barang or k.get_kondisi_display()
            obj.lokasi_barang = obj.lokasi_barang or str(k.unit_kerja or '')
        elif obj.jenis_aset == 'RUMAH_NEGARA' and obj.rumah_negara:
            r = obj.rumah_negara
            obj.kode_barang = obj.kode_barang or r.kode_barang
            obj.nup = obj.nup or r.nup
            obj.nama_barang = obj.nama_barang or f'{r.kode_rumah} - {r.nama_rumah}'
            obj.nilai_perolehan = obj.nilai_perolehan or r.nilai_perolehan
            obj.kondisi_barang = obj.kondisi_barang or r.get_kondisi_display()
            obj.lokasi_barang = obj.lokasi_barang or r.alamat
        elif obj.jenis_aset == 'TANAH_NEGARA' and obj.tanah_negara:
            t = obj.tanah_negara
            obj.kode_barang = obj.kode_barang or t.kode_barang
            obj.nup = obj.nup or t.nup
            obj.nama_barang = obj.nama_barang or f'{t.kode_tanah} - {t.nama_tanah}'
            obj.nilai_perolehan = obj.nilai_perolehan or t.nilai_perolehan
            obj.kondisi_barang = obj.kondisi_barang or t.get_status_tanah_display()
            obj.lokasi_barang = obj.lokasi_barang or t.alamat


class PermohonanPenghapusanUpdateView(PermohonanPenghapusanAccessMixin, UpdateView):
    model = PermohonanPenghapusanBMN
    form_class = PermohonanPenghapusanBMNForm
    template_name = 'penghapusan/form.html'
    success_url = reverse_lazy('penghapusan:list')

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        if is_sekretaris_jenderal(request.user) and not request.user.is_superuser:
            raise PermissionDenied('Role Sekjen memproses penghapusan melalui menu Penetapan SK, bukan edit umum.')
        # Perbaikan 2026-06-16:
        # Role Pengelola BMN yang masih berada dalam scope datanya boleh mengedit
        # Permohonan Penghapusan BMN dari menu Permohonan. Sebelumnya status
        # DIVERIFIKASI_BIRO/DIAJUKAN_SEKJEN memicu 403 walaupun tombol Edit tampil.
        # Keamanan scope tetap dijaga oleh get_object() dari get_scoped_queryset().
        if self.object.status_norm in ['SK_PENGHAPUSAN_TERBIT', 'SELESAI'] and not (is_biro_umum_user(request.user) or request.user.is_superuser):
            raise PermissionDenied('Usulan yang sudah selesai/SK terbit tidak dapat diedit oleh Pengelola BMN.')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        user = self.request.user
        form.instance.diperbarui_oleh = user
        if is_global_bmn_scope_user(user):
            if getattr(form.instance, 'status_norm', form.instance.status) in ['DIVERIFIKASI_BIRO_UMUM', 'DIAJUKAN_KE_DIRJEN_REHSOS', 'DISETUJUI_DIRJEN_REHSOS', 'DIAJUKAN_KE_SEKJEN', 'SK_PENGHAPUSAN_TERBIT', 'SELESAI'] and not form.instance.tanggal_verifikasi:
                form.instance.tanggal_verifikasi = timezone.now().date()
                form.instance.diverifikasi_oleh = user
        else:
            # Saat Pengelola BMN mengedit usulan, jangan paksa status kembali
            # menjadi DIAJUKAN karena bisa mengacaukan alur verifikasi Biro Umum.
            # Status lama dipertahankan; khusus DRAFT/PERLU_PERBAIKAN boleh naik
            # menjadi DIAJUKAN agar usulan kembali masuk antrian.
            old_status = getattr(self.object, 'status', None) if getattr(self, 'object', None) else None
            if old_status in ['DRAFT', 'PERLU_PERBAIKAN']:
                form.instance.status = 'MENUNGGU_VERIFIKASI_BIRO_UMUM'
            elif old_status:
                form.instance.status = old_status
            else:
                form.instance.status = 'MENUNGGU_VERIFIKASI_BIRO_UMUM'
        response = super().form_valid(form)
        _save_foto_kondisi_files(self.request, self.object)
        messages.success(self.request, 'Permohonan penghapusan BMN berhasil diperbarui.')
        return response


class PermohonanPenghapusanDetailView(PermohonanPenghapusanAccessMixin, DetailView):
    model = PermohonanPenghapusanBMN
    template_name = 'penghapusan/detail.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['foto_kondisi_list'] = self.object.foto_kondisi_list.all()
        ctx['can_verifikasi_penghapusan'] = (is_biro_umum_user(self.request.user) or self.request.user.is_superuser) and self.object.status_norm in ['DIAJUKAN_UNIT_KERJA', 'MENUNGGU_VERIFIKASI_BIRO_UMUM', 'DIVERIFIKASI_BIRO_UMUM', 'PERLU_PERBAIKAN']
        ctx['can_persetujuan_dirjen_rehsos'] = (is_dirjen_rehsos(self.request.user) or self.request.user.is_superuser) and self.object.status_norm in ['DIAJUKAN_KE_DIRJEN_REHSOS', 'DISETUJUI_DIRJEN_REHSOS', 'DITOLAK_DIRJEN_REHSOS']
        ctx['can_penetapan_penghapusan'] = (is_sekretaris_jenderal(self.request.user) or self.request.user.is_superuser) and self.object.status_norm in ['DIAJUKAN_KE_SEKJEN', 'DISETUJUI_DIRJEN_REHSOS', 'DITOLAK_SEKJEN', 'SK_PENGHAPUSAN_TERBIT']
        ctx['back_url_name'] = self.request.GET.get('from') or ('penghapusan:persetujuan_sekjen' if is_sekretaris_jenderal(self.request.user) else ('penghapusan:persetujuan_dirjen_rehsos' if is_dirjen_rehsos(self.request.user) else ('penghapusan:verifikasi' if is_biro_umum_user(self.request.user) else 'penghapusan:list')))
        ctx['alur_nota_penghapusan'] = get_alur_nota_penghapusan(self.object)
        return ctx



def _norm_header(value):
    return str(value or '').strip().lower().replace(' ', '_').replace('-', '_')


def _dec(value):
    if value in [None, '']:
        return Decimal('0')
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    cleaned = str(value).strip().replace('Rp', '').replace('rp', '').replace(' ', '')
    cleaned = cleaned.replace('.', '').replace(',', '.') if ',' in cleaned else cleaned.replace(',', '')
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return Decimal('0')

def _unit_text(unit):
    if not unit:
        return ''
    parts = []
    for field in ['nama_unit', 'nama', 'nama_unit_kerja', 'unit_eselon_i', 'nama_eselon_i', 'jenis_unit']:
        value = getattr(unit, field, None)
        if value:
            parts.append(str(value))
    return ' '.join(parts).upper()


def get_alur_nota_penghapusan(obj):
    """Menentukan tujuan nota dinas penghapusan sesuai aturan bisnis terakhir.

    - Sentra di bawah Ditjen Rehsos: nota dinas Sentra ditujukan kepada Dirjen Rehsos.
    - Balai: nota dinas langsung ditujukan kepada Sekretaris Jenderal.
    - Unit pusat: pengusul/penandatangan nota adalah Sekretaris Unit Eselon I kepada Sekretaris Jenderal.
    """
    unit = getattr(obj, 'unit_kerja', None)
    text = _unit_text(unit)
    if 'SENTRA' in text and ('REHABILITASI SOSIAL' in text or 'REHSOS' in text):
        return {
            'kategori': 'SENTRA_REHSOS',
            'tujuan': 'Direktur Jenderal Rehabilitasi Sosial, lalu diteruskan kepada Sekretaris Jenderal',
            'dari': 'Kepala Sentra → Dirjen Rehabilitasi Sosial',
            'keterangan': 'Usulan penghapusan dari Sentra di bawah Ditjen Rehabilitasi Sosial diproses berjenjang: Kepala Sentra mengajukan nota dinas kepada Dirjen Rehabilitasi Sosial, kemudian Dirjen Rehabilitasi Sosial meneruskan nota dinas kepada Sekretaris Jenderal untuk penetapan SK.',
        }
    if 'BALAI' in text:
        return {
            'kategori': 'BALAI',
            'tujuan': 'Sekretaris Jenderal',
            'dari': 'Kepala Balai',
            'keterangan': 'Usulan penghapusan dari Balai: nota dinas langsung ditujukan kepada Sekretaris Jenderal setelah diverifikasi/diadministrasikan sesuai alur Biro Umum.',
        }
    # default unit pusat/eselon I
    return {
        'kategori': 'PUSAT_ESELON_I',
        'tujuan': 'Sekretaris Jenderal',
        'dari': 'Sekretaris Unit Eselon I',
        'keterangan': 'Usulan penghapusan dari unit pusat: yang mengajukan adalah Sekretaris Unit Eselon I kepada Sekretaris Jenderal.',
    }



class DownloadTemplateBarangPenghapusanView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Barang Penghapusan'
        headers = ['no', 'kode_barang', 'nup', 'nama_barang', 'jenis_aset', 'kuantitas', 'nilai_perolehan', 'kondisi_barang', 'lokasi_barang', 'alasan_penghapusan', 'keterangan']
        ws.append(headers)
        ws.append([1, '3100102002', '9471', 'Laptop rusak berat', 'LAINNYA', 1, 9183030, 'Rusak Berat', 'Gudang Satker', 'RUSAK_BERAT', 'Contoh barang yang diusulkan hapus'])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(max(len(str(col[0].value or '')) + 4, 14), 36)
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        response = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="template_import_barang_penghapusan.xlsx"'
        return response


class ImportBarangPenghapusanView(PermohonanPenghapusanAccessMixin, View):
    template_name = 'penghapusan/import_barang.html'

    def dispatch(self, request, *args, **kwargs):
        self.object = get_object_or_404(self.get_scoped_queryset(), pk=kwargs['pk'])
        if is_sekretaris_jenderal(request.user) and not request.user.is_superuser:
            raise PermissionDenied('Role Sekjen hanya dapat melihat/menetapkan. Import barang penghapusan tidak diperbolehkan.')
        if (not is_biro_umum_user(request.user) and not request.user.is_superuser
                and self.object.status_norm in ['SK_PENGHAPUSAN_TERBIT', 'SELESAI']):
            raise PermissionDenied('Import barang penghapusan tidak dapat dilakukan jika usulan sudah selesai/SK terbit.')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name, {'object': self.object, 'form': ImportBarangPenghapusanForm()})

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        form = ImportBarangPenghapusanForm(request.POST, request.FILES)
        if not form.is_valid():
            return render(request, self.template_name, {'object': self.object, 'form': form})
        if form.cleaned_data.get('replace_existing'):
            self.object.detail_barang.all().delete()
        wb = load_workbook(form.cleaned_data['file_excel'], data_only=True)
        ws = wb.active
        headers = [_norm_header(c.value) for c in ws[1]]
        created = 0
        for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = dict(zip(headers, row))
            if not any(data.values()):
                continue
            def pick(*keys):
                for key in keys:
                    if key in data and data[key] not in [None, '']:
                        return data[key]
                return ''
            nama = str(pick('nama_barang', 'barang', 'nama') or '').strip()
            if not nama:
                continue
            jenis = str(pick('jenis_aset') or self.object.jenis_aset or 'LAINNYA').strip().upper()
            if jenis not in dict(PermohonanPenghapusanBMN.JENIS_ASET):
                jenis = 'LAINNYA'
            alasan = str(pick('alasan_penghapusan') or self.object.alasan_penghapusan or 'RUSAK_BERAT').strip().upper()
            if alasan not in dict(PermohonanPenghapusanBMN.ALASAN_PENGHAPUSAN):
                alasan = 'LAINNYA'
            BarangPenghapusanBMN.objects.create(
                permohonan=self.object,
                nomor_urut=int(pick('no', 'nomor_urut') or created + 1),
                kode_barang=str(pick('kode_barang') or ''),
                nup=str(pick('nup') or ''),
                nama_barang=nama,
                jenis_aset=jenis,
                kuantitas=int(pick('kuantitas', 'jumlah', 'qty') or 1),
                nilai_perolehan=_dec(pick('nilai_perolehan', 'nilai')),
                kondisi_barang=str(pick('kondisi_barang', 'kondisi') or ''),
                lokasi_barang=str(pick('lokasi_barang', 'lokasi') or ''),
                alasan_penghapusan=alasan,
                keterangan=str(pick('keterangan') or ''),
            )
            created += 1
        # Update ringkasan utama dari detail pertama/total agar daftar tetap informatif.
        first = self.object.detail_barang.first()
        if first:
            self.object.nama_barang = first.nama_barang if created == 1 else f'{created} unit barang BMN yang diusulkan hapus'
            self.object.nilai_perolehan = sum([x.nilai_perolehan * x.kuantitas for x in self.object.detail_barang.all()], Decimal('0'))
            self.object.kode_barang = first.kode_barang
            self.object.nup = first.nup
            self.object.save(update_fields=['nama_barang', 'nilai_perolehan', 'kode_barang', 'nup', 'updated_at'])
        messages.success(request, f'{created} baris barang penghapusan berhasil diimport.')
        return redirect('penghapusan:detail', pk=self.object.pk)

class ProsesPenghapusanView(PermohonanPenghapusanAccessMixin, View):
    """Aksi workflow Penghapusan BMN, disamakan dengan pola PSP.

    Alur umum: Unit Kerja -> Verifikasi Biro Umum -> Penetapan Sekjen -> SK Terbit.
    Khusus Sentra Rehsos: Unit Kerja -> Verifikasi Biro Umum -> Dirjen Rehsos -> Sekjen -> SK Terbit.
    """
    def _is_sentra_rehsos(self, obj):
        return get_alur_nota_penghapusan(obj).get('kategori') == 'SENTRA_REHSOS'

    def post(self, request, *args, **kwargs):
        obj = get_object_or_404(PermohonanPenghapusanBMN.objects.select_related('unit_kerja', 'pemohon'), pk=kwargs['pk'])
        action = request.POST.get('action')
        catatan = (request.POST.get('catatan') or '').strip()
        today = timezone.now().date()

        if action in ['verifikasi_biro', 'kembalikan_biro', 'teruskan_sekjen']:
            if not (is_biro_umum_user(request.user) or request.user.is_superuser):
                raise PermissionDenied('Hanya role Biro Umum yang dapat memverifikasi usulan penghapusan.')
            obj.diverifikasi_oleh = request.user
            obj.tanggal_verifikasi = obj.tanggal_verifikasi or today
            if catatan:
                obj.catatan_biro_umum = catatan
            if request.FILES.get('dokumen_penghapusan_final'):
                obj.dokumen_persetujuan = request.FILES['dokumen_penghapusan_final']
            if action == 'verifikasi_biro':
                obj.status = 'DIVERIFIKASI_BIRO_UMUM'
                messages.success(request, 'Usulan penghapusan berhasil ditandai diverifikasi Biro Umum.')
            elif action == 'kembalikan_biro':
                obj.status = 'PERLU_PERBAIKAN'
                messages.success(request, 'Usulan penghapusan dikembalikan ke unit kerja untuk perbaikan.')
            else:
                if not obj.dokumen_persetujuan:
                    messages.error(request, 'Upload Dokumen Penghapusan SIKARIS Final/Gabungan PDF terlebih dahulu sebelum diteruskan.')
                    return redirect('penghapusan:detail', pk=obj.pk)
                if self._is_sentra_rehsos(obj):
                    obj.status = 'DIAJUKAN_KE_DIRJEN_REHSOS'
                    messages.success(request, 'Usulan Sentra Rehsos diteruskan ke Dirjen Rehabilitasi Sosial terlebih dahulu sebelum ke Sekjen.')
                else:
                    obj.status = 'DIAJUKAN_KE_SEKJEN'
                    messages.success(request, 'Usulan penghapusan berhasil diteruskan ke Sekjen untuk penetapan SK.')
            obj.diperbarui_oleh = request.user
            obj.save()
            return redirect('penghapusan:detail', pk=obj.pk)

        if action in ['setujui_dirjen_rehsos', 'tolak_dirjen_rehsos']:
            if not (is_dirjen_rehsos(request.user) or request.user.is_superuser):
                raise PermissionDenied('Hanya role Dirjen Rehsos yang dapat memproses usulan Sentra Rehsos.')
            if not self._is_sentra_rehsos(obj):
                raise PermissionDenied('Persetujuan Dirjen Rehsos hanya untuk usulan dari Sentra di bawah Ditjen Rehsos.')
            if action == 'setujui_dirjen_rehsos':
                obj.status = 'DIAJUKAN_KE_SEKJEN'
                messages.success(request, 'Usulan penghapusan Sentra Rehsos disetujui Dirjen Rehsos dan diteruskan ke Sekjen.')
            else:
                obj.status = 'DITOLAK_DIRJEN_REHSOS'
                if catatan:
                    obj.catatan_unit = catatan
                messages.success(request, 'Usulan penghapusan Sentra Rehsos ditolak/dikembalikan oleh Dirjen Rehsos.')
            obj.diperbarui_oleh = request.user
            obj.save()
            return redirect('penghapusan:detail', pk=obj.pk)

        if action in ['tetapkan_sekjen', 'tolak_sekjen']:
            if not (is_sekretaris_jenderal(request.user) or request.user.is_superuser):
                raise PermissionDenied('Hanya role Sekjen yang dapat menetapkan SK Penghapusan BMN.')
            if action == 'tetapkan_sekjen':
                obj.status = 'SK_PENGHAPUSAN_TERBIT'
                obj.tanggal_persetujuan = obj.tanggal_persetujuan or today
                obj.tanggal_sk_penghapusan = obj.tanggal_sk_penghapusan or today
                if not obj.nomor_sk_penghapusan:
                    obj.nomor_sk_penghapusan = f'SK-PENGHAPUSAN/{today:%Y}/{obj.pk:05d}'
                if request.FILES.get('dokumen_sk_penghapusan'):
                    obj.dokumen_sk_penghapusan = request.FILES['dokumen_sk_penghapusan']
                messages.success(request, 'SK Penghapusan BMN berhasil ditetapkan oleh Sekjen.')
            else:
                obj.status = 'DITOLAK_SEKJEN'
                if catatan:
                    obj.catatan_unit = catatan
                messages.success(request, 'Usulan penghapusan ditolak oleh Sekjen.')
            obj.diperbarui_oleh = request.user
            obj.save()
            return redirect('penghapusan:detail', pk=obj.pk)

        raise PermissionDenied('Aksi proses penghapusan tidak dikenal.')


class PermohonanPenghapusanDeleteView(PermohonanPenghapusanAccessMixin, DeleteView):
    """Hapus Permohonan Penghapusan BMN.

    Catatan: view ini sengaja tidak memakai form bawaan DeleteView/SafeDeleteMixin
    karena AccessMixin menambahkan `user` pada form kwargs. Pada DeleteView hal itu
    dapat membuat tombol submit tidak memproses dengan benar atau memunculkan error
    `__init__() got an unexpected keyword argument user`.
    """
    model = PermohonanPenghapusanBMN
    template_name = 'includes/confirm_delete.html'
    success_url = reverse_lazy('penghapusan:list')
    success_message = 'Permohonan penghapusan BMN berhasil dihapus.'

    def get_form_kwargs(self):
        # Jangan kirim kwargs apa pun ke form delete bawaan Django.
        return {}

    def _check_delete_permission(self, obj):
        if is_sekretaris_jenderal(self.request.user) and not self.request.user.is_superuser:
            raise PermissionDenied('Role Sekjen tidak diperbolehkan menghapus usulan penghapusan.')

        # Admin/Biro Umum dapat menghapus sesuai kewenangan global.
        # Pengelola BMN hanya boleh menghapus data dalam scope-nya dan belum final.
        if not is_global_bmn_scope_user(self.request.user):
            allowed_status = [
                'DRAFT',
                'DIAJUKAN_UNIT_KERJA',
                'MENUNGGU_VERIFIKASI_BIRO_UMUM',
                'PERLU_PERBAIKAN',
            ]
            if obj.status_norm not in allowed_status:
                raise PermissionDenied('Usulan yang sudah diproses lanjut tidak dapat dihapus oleh unit kerja.')

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        self._check_delete_permission(self.object)
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        # Proses hapus langsung dari POST agar tombol Ya, Hapus pasti bekerja.
        self.object = self.get_object()
        self._check_delete_permission(self.object)
        try:
            self.object.delete()
            messages.success(request, self.success_message)
        except ProtectedError:
            messages.error(request, 'Data tidak dapat dihapus karena masih digunakan oleh data lain.')
        return redirect(self.get_success_url())

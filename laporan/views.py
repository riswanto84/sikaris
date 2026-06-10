from io import BytesIO
from textwrap import wrap

from django.contrib.auth.decorators import login_required, user_passes_test
from django.conf import settings
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from svglib.svglib import svg2rlg
except Exception:  # pragma: no cover - fallback bila dependency PDF belum terpasang
    colors = None
    TA_CENTER = TA_LEFT = None
    A4 = None
    ParagraphStyle = getSampleStyleSheet = None
    mm = None
    SimpleDocTemplate = Paragraph = Spacer = Table = TableStyle = None
    svg2rlg = None

from core.roles import can_view_reports, can_manage_master
from master.models import Kendaraan, RumahDinas, Pegawai
from kendaraan.models import SIPKendaraan, ServiceKendaraan, RiwayatKondisiKendaraan
from rumah_dinas.models import SIPRumahDinas


# -----------------------------------------------------------------------------
# Helper umum export
# -----------------------------------------------------------------------------
def _today_label():
    return timezone.localtime().strftime('%Y%m%d_%H%M')


def _display(obj, field_name, default='-'):
    """Ambil get_FIELD_display() kalau ada; kalau tidak, ambil value biasa."""
    if obj is None:
        return default
    display_func = getattr(obj, f'get_{field_name}_display', None)
    if callable(display_func):
        value = display_func()
    else:
        value = getattr(obj, field_name, None)
    return value if value not in [None, ''] else default


def _text(value, default='-'):
    if value is None or value == '':
        return default
    return str(value)


def _date(value):
    return value.strftime('%d-%m-%Y') if value else '-'


def _money(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0


def _rupiah_text(value):
    return f"Rp {float(value or 0):,.0f}".replace(',', '.')


def _autosize_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 45)


def _style_sheet(ws):
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    _autosize_columns(ws)


def _excel_response(wb, filename):
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _pegawai_identity(pegawai):
    if not pegawai:
        return '-'
    nip = getattr(pegawai, 'nip', None)
    return f"{pegawai.nama} ({nip})" if nip else pegawai.nama


def _format_sip_history(sip):
    pegawai = getattr(sip, 'pegawai', None)
    return (
        f"{_pegawai_identity(pegawai)} | SIP {sip.nomor_sip} | "
        f"{_date(sip.tanggal_mulai)} s.d. {_date(sip.tanggal_akhir)} | {_display(sip, 'status')}"
    )


def _get_previous_sips(all_sips, latest_sip):
    previous = []
    for sip in all_sips:
        if latest_sip and sip.pk == latest_sip.pk:
            continue
        previous.append(sip)
    return previous


def _join_previous_users(previous_sips):
    if not previous_sips:
        return '-'
    return '; '.join(_format_sip_history(sip) for sip in previous_sips)


def _logo_path():
    candidates = [
        settings.BASE_DIR / 'static' / 'img' / 'logo-kemensos.svg',
        settings.BASE_DIR / 'logo-kemensos.svg',
        settings.BASE_DIR / 'staticfiles' / 'img' / 'logo-kemensos.svg',
    ]
    for path in candidates:
        if path.exists():
            return str(path)
    return None


# -----------------------------------------------------------------------------
# PDF sederhana tanpa dependency tambahan.
# Cocok agar export PDF tetap berjalan di project lokal tanpa install reportlab.
# -----------------------------------------------------------------------------
def _pdf_escape(text):
    text = str(text).replace('—', '-').replace('–', '-').replace('“', '"').replace('”', '"')
    return text.replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')


def _build_simple_pdf(lines, title='Laporan SIKARIS'):
    """Buat PDF dengan header logo Kemensos.

    Menggunakan ReportLab agar header dapat menampilkan logo. Bila dependency
    belum terpasang, fungsi akan fallback ke PDF teks sederhana.
    """
    if not all([SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle]):
        return _build_fallback_pdf(lines, title)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=32 * mm,
        bottomMargin=15 * mm,
        title=title,
    )

    styles = getSampleStyleSheet()
    body_style = ParagraphStyle(
        'SIKARISBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=11,
        spaceAfter=2,
        alignment=TA_LEFT,
    )
    title_style = ParagraphStyle(
        'SIKARISTitle',
        parent=styles['Title'],
        fontName='Helvetica-Bold',
        fontSize=13,
        leading=16,
        alignment=TA_CENTER,
        spaceAfter=6,
    )

    story = []
    for idx, line in enumerate(lines):
        safe_line = _text(line, '')
        if idx == 0:
            story.append(Paragraph(safe_line, title_style))
        elif safe_line == '':
            story.append(Spacer(1, 4))
        else:
            story.append(Paragraph(safe_line.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;'), body_style))

    logo = _logo_path()

    def header_footer(canvas, doc_obj):
        canvas.saveState()
        width, height = A4
        y = height - 20 * mm
        if logo and svg2rlg:
            try:
                drawing = svg2rlg(logo)
                if drawing:
                    target_h = 16 * mm
                    scale = target_h / drawing.height
                    drawing.scale(scale, scale)
                    drawing.drawOn(canvas, 15 * mm, height - 22 * mm)
            except Exception:
                canvas.setFont('Helvetica-Bold', 8)
                canvas.drawString(15 * mm, y, 'Kemensos RI')
        else:
            canvas.setFont('Helvetica-Bold', 8)
            canvas.drawString(15 * mm, y, 'Kemensos RI')
        canvas.setFont('Helvetica-Bold', 10)
        canvas.drawCentredString(width / 2, height - 14 * mm, 'KEMENTERIAN SOSIAL REPUBLIK INDONESIA')
        canvas.setFont('Helvetica', 8)
        canvas.drawCentredString(width / 2, height - 19 * mm, 'SIKARIS - Sistem Informasi Kendaraan dan Rumah Dinas')
        canvas.setStrokeColor(colors.HexColor('#1F4E79'))
        canvas.line(15 * mm, height - 26 * mm, width - 15 * mm, height - 26 * mm)
        canvas.setFont('Helvetica', 7)
        canvas.drawRightString(width - 15 * mm, 9 * mm, f'Halaman {doc_obj.page}')
        canvas.restoreState()

    doc.build(story, onFirstPage=header_footer, onLaterPages=header_footer)
    buffer.seek(0)
    return buffer.read()


def _build_fallback_pdf(lines, title='Laporan SIKARIS'):
    # Ukuran A4 portrait: 595 x 842 pt
    page_width = 595
    page_height = 842
    margin_left = 36
    start_y = 800
    line_height = 13
    font_size = 9
    max_chars = 96

    header_lines = [
        'KEMENTERIAN SOSIAL REPUBLIK INDONESIA',
        'SIKARIS - Sistem Informasi Kendaraan dan Rumah Dinas',
        '',
    ]
    prepared_lines = []
    for line in header_lines + list(lines):
        if line == '':
            prepared_lines.append('')
            continue
        for part in wrap(str(line), width=max_chars, replace_whitespace=False, drop_whitespace=False) or ['']:
            prepared_lines.append(part)

    pages = []
    current = []
    max_lines_per_page = int((start_y - 40) / line_height)
    for line in prepared_lines:
        current.append(line)
        if len(current) >= max_lines_per_page:
            pages.append(current)
            current = []
    if current:
        pages.append(current)

    objects = []
    objects.append('<< /Type /Catalog /Pages 2 0 R >>')
    page_kids_placeholder_index = 1
    objects.append('')
    objects.append('<< /Type /Font /Subtype /Type1 /BaseFont /Courier >>')

    page_object_numbers = []
    for page_lines in pages:
        content_commands = ['BT', f'/F1 {font_size} Tf', f'{margin_left} {start_y} Td']
        first = True
        for line in page_lines:
            if not first:
                content_commands.append(f'0 -{line_height} Td')
            content_commands.append(f'({_pdf_escape(line)}) Tj')
            first = False
        content_commands.append('ET')
        content = '\n'.join(content_commands)
        stream = f'<< /Length {len(content.encode("latin-1", errors="replace"))} >>\nstream\n{content}\nendstream'
        content_obj_num = len(objects) + 1
        objects.append(stream)
        page_obj_num = len(objects) + 1
        page_object_numbers.append(page_obj_num)
        objects.append(
            f'<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_width} {page_height}] '
            f'/Resources << /Font << /F1 3 0 R >> >> /Contents {content_obj_num} 0 R >>'
        )

    kids = ' '.join(f'{num} 0 R' for num in page_object_numbers)
    objects[page_kids_placeholder_index] = f'<< /Type /Pages /Kids [{kids}] /Count {len(page_object_numbers)} >>'

    pdf = bytearray(b'%PDF-1.4\n')
    offsets = [0]
    for idx, obj in enumerate(objects, start=1):
        offsets.append(len(pdf))
        pdf.extend(f'{idx} 0 obj\n'.encode('ascii'))
        pdf.extend(obj.encode('latin-1', errors='replace'))
        pdf.extend(b'\nendobj\n')
    xref_offset = len(pdf)
    pdf.extend(f'xref\n0 {len(objects)+1}\n'.encode('ascii'))
    pdf.extend(b'0000000000 65535 f \n')
    for offset in offsets[1:]:
        pdf.extend(f'{offset:010d} 00000 n \n'.encode('ascii'))
    pdf.extend(
        f'trailer\n<< /Size {len(objects)+1} /Root 1 0 R /Title ({_pdf_escape(title)}) >>\n'
        f'startxref\n{xref_offset}\n%%EOF'.encode('latin-1', errors='replace')
    )
    return bytes(pdf)


def _pdf_response(lines, filename, title='Laporan SIKARIS'):
    response = HttpResponse(_build_simple_pdf(lines, title), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@user_passes_test(can_view_reports, login_url='login')
def laporan_index(request):
    return render(request, 'laporan/index.html', {
        'kendaraan': Kendaraan.objects.count(),
        'rumah': RumahDinas.objects.count(),
        'pegawai': Pegawai.objects.count(),
        'sip_kendaraan': SIPKendaraan.objects.count(),
        'sip_rumah': SIPRumahDinas.objects.count(),
        'service': ServiceKendaraan.objects.count(),
    })


# -----------------------------------------------------------------------------
# Export Laporan Kendaraan
# -----------------------------------------------------------------------------
def _get_latest_sip_kendaraan(kendaraan):
    return kendaraan.sip_kendaraan.all().order_by('-tanggal_sip').first()


@login_required
@user_passes_test(can_view_reports, login_url='login')
def export_kendaraan_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Kendaraan dan SIP'

    headers = [
        'Kode Kendaraan', 'Nomor Polisi', 'Merek', 'Tipe', 'Jenis', 'Tahun', 'Unit Kerja',
        'Pengguna Saat Ini', 'Kondisi', 'Status Pemanfaatan', 'KM Terakhir', 'Nilai Perolehan',
        'Nomor SIP Terakhir', 'Tanggal SIP', 'Pemakai SIP', 'Periode SIP', 'Status SIP',
        'Lokasi Penggunaan', 'Catatan SIP', 'Riwayat Pengguna Sebelumnya'
    ]
    ws.append(headers)

    kendaraan_qs = Kendaraan.objects.select_related('unit_kerja', 'pengguna').prefetch_related(
        'sip_kendaraan__pegawai', 'service', 'riwayat_kondisi'
    ).order_by('nomor_polisi')

    for k in kendaraan_qs:
        sip = _get_latest_sip_kendaraan(k)
        previous_sips = _get_previous_sips(
            list(k.sip_kendaraan.all().order_by('-tanggal_mulai', '-tanggal_sip')),
            sip
        )
        ws.append([
            k.kode_kendaraan, k.nomor_polisi, k.merek, _text(k.tipe), _display(k, 'jenis_kendaraan'),
            _text(k.tahun_pembuatan), _text(k.unit_kerja), _text(k.pengguna), _display(k, 'kondisi'),
            _display(k, 'status_pemanfaatan'), k.kilometer_terakhir, _money(k.nilai_perolehan),
            _text(sip.nomor_sip if sip else None), _date(sip.tanggal_sip if sip else None),
            _text(sip.pegawai if sip else None),
            f"{_date(sip.tanggal_mulai)} s.d. {_date(sip.tanggal_akhir)}" if sip else '-',
            _display(sip, 'status') if sip else '-', _text(sip.lokasi_penggunaan if sip else None),
            _text(sip.catatan if sip else None), _join_previous_users(previous_sips),
        ])
    _style_sheet(ws)

    ws_sip = wb.create_sheet('Detail SIP Kendaraan')
    ws_sip.append([
        'Nomor SIP', 'Tanggal SIP', 'Nomor Polisi', 'Kendaraan', 'Pemakai', 'NIP Pemakai',
        'Tanggal Mulai', 'Tanggal Akhir', 'Jenis Pemakaian', 'Tujuan Pemakaian',
        'Lokasi Penggunaan', 'Status SIP', 'Pejabat Penandatangan', 'Catatan'
    ])
    for sip in SIPKendaraan.objects.select_related('kendaraan', 'pegawai').order_by('-tanggal_sip'):
        ws_sip.append([
            sip.nomor_sip, _date(sip.tanggal_sip), sip.kendaraan.nomor_polisi,
            _text(sip.kendaraan), sip.pegawai.nama, sip.pegawai.nip,
            _date(sip.tanggal_mulai), _date(sip.tanggal_akhir), _display(sip, 'jenis_pemakaian'),
            _text(sip.tujuan_pemakaian), _text(sip.lokasi_penggunaan), _display(sip, 'status'),
            _text(sip.pejabat_penandatangan), _text(sip.catatan),
        ])
    _style_sheet(ws_sip)

    ws_pengguna = wb.create_sheet('Riwayat Pengguna Kendaraan')
    ws_pengguna.append([
        'Nomor Polisi', 'Kendaraan', 'Nomor SIP', 'Tanggal SIP', 'Nama Pegawai', 'NIP',
        'Jabatan', 'Unit Kerja', 'Tanggal Mulai', 'Tanggal Akhir', 'Status SIP',
        'Jenis Pemakaian', 'Lokasi Penggunaan', 'Tujuan Pemakaian'
    ])
    for sip in SIPKendaraan.objects.select_related(
        'kendaraan', 'pegawai__unit_kerja'
    ).order_by('kendaraan__nomor_polisi', '-tanggal_mulai', '-tanggal_sip'):
        ws_pengguna.append([
            sip.kendaraan.nomor_polisi, _text(sip.kendaraan), sip.nomor_sip, _date(sip.tanggal_sip),
            sip.pegawai.nama, sip.pegawai.nip, _text(sip.pegawai.jabatan), _text(sip.pegawai.unit_kerja),
            _date(sip.tanggal_mulai), _date(sip.tanggal_akhir), _display(sip, 'status'),
            _display(sip, 'jenis_pemakaian'), _text(sip.lokasi_penggunaan), _text(sip.tujuan_pemakaian),
        ])
    _style_sheet(ws_pengguna)

    ws_service = wb.create_sheet('Riwayat Service')
    ws_service.append([
        'Tanggal Service', 'Nomor Polisi', 'Kendaraan', 'Jenis Service', 'Kilometer', 'Bengkel',
        'Uraian Pekerjaan', 'Sparepart Diganti', 'Biaya Jasa', 'Biaya Sparepart', 'Total Biaya',
        'Kondisi Sebelum', 'Kondisi Sesudah'
    ])
    for s in ServiceKendaraan.objects.select_related('kendaraan').order_by('-tanggal_service'):
        ws_service.append([
            _date(s.tanggal_service), s.kendaraan.nomor_polisi, _text(s.kendaraan),
            _display(s, 'jenis_service'), _text(s.kilometer), _text(s.bengkel),
            _text(s.uraian_pekerjaan), _text(s.sparepart_diganti), _money(s.biaya_jasa),
            _money(s.biaya_sparepart), _money(s.total_biaya), _display(s, 'kondisi_sebelum'),
            _display(s, 'kondisi_sesudah'),
        ])
    _style_sheet(ws_service)

    ws_kondisi = wb.create_sheet('Riwayat Kondisi Kendaraan')
    ws_kondisi.append([
        'Tanggal', 'Nomor Polisi', 'Kendaraan', 'Kondisi', 'Uraian Kondisi', 'Dicatat Oleh'
    ])
    for rk in RiwayatKondisiKendaraan.objects.select_related('kendaraan', 'dicatat_oleh').order_by('-tanggal'):
        ws_kondisi.append([
            _date(rk.tanggal), rk.kendaraan.nomor_polisi, _text(rk.kendaraan),
            _display(rk, 'kondisi'), _text(rk.uraian_kondisi), _text(rk.dicatat_oleh),
        ])
    _style_sheet(ws_kondisi)

    return _excel_response(wb, f'laporan_kendaraan_sip_service_{_today_label()}.xlsx')


@login_required
@user_passes_test(can_view_reports, login_url='login')
def export_kendaraan_pdf(request):
    lines = []
    lines.append('LAPORAN KENDARAAN, SIP, KONDISI, DAN RIWAYAT SERVICE')
    lines.append(f'Dicetak: {timezone.localtime().strftime("%d-%m-%Y %H:%M") }')
    lines.append('')

    kendaraan_qs = Kendaraan.objects.select_related('unit_kerja', 'pengguna').prefetch_related(
        'sip_kendaraan__pegawai', 'service', 'riwayat_kondisi'
    ).order_by('nomor_polisi')

    for no, k in enumerate(kendaraan_qs, start=1):
        sip = _get_latest_sip_kendaraan(k)
        lines.append(f'{no}. {k.nomor_polisi} - {k.merek} {_text(k.tipe, "")}'.strip())
        lines.append(f'   Kode: {k.kode_kendaraan} | Jenis: {_display(k, "jenis_kendaraan")} | Tahun: {_text(k.tahun_pembuatan)}')
        lines.append(f'   Unit Kerja: {_text(k.unit_kerja)} | Pengguna: {_text(k.pengguna)}')
        lines.append(f'   Kondisi: {_display(k, "kondisi")} | Status: {_display(k, "status_pemanfaatan")} | KM: {k.kilometer_terakhir}')
        if sip:
            lines.append(f'   SIP terakhir: {sip.nomor_sip} ({_display(sip, "status")}), tanggal {_date(sip.tanggal_sip)}')
            lines.append(f'   Pemakai SIP: {_text(sip.pegawai)} | Periode: {_date(sip.tanggal_mulai)} s.d. {_date(sip.tanggal_akhir)}')
            lines.append(f'   Lokasi/Tujuan: {_text(sip.lokasi_penggunaan)} / {_text(sip.tujuan_pemakaian)}')
        else:
            lines.append('   SIP terakhir: -')

        previous_sips = _get_previous_sips(
            list(k.sip_kendaraan.all().order_by('-tanggal_mulai', '-tanggal_sip')),
            sip
        )
        if previous_sips:
            lines.append('   Riwayat pengguna sebelumnya:')
            for prev_sip in previous_sips[:10]:
                lines.append(f'   - {_format_sip_history(prev_sip)}')
        else:
            lines.append('   Riwayat pengguna sebelumnya: -')

        kondisi_logs = k.riwayat_kondisi.all().order_by('-tanggal')[:5]
        if kondisi_logs:
            lines.append('   Riwayat kondisi terakhir:')
            for rk in kondisi_logs:
                lines.append(f'   - {_date(rk.tanggal)} | {_display(rk, "kondisi")} | {_text(rk.uraian_kondisi)}')
        else:
            lines.append('   Riwayat kondisi: -')

        services = k.service.all().order_by('-tanggal_service')[:5]
        if services:
            lines.append('   Riwayat service terakhir:')
            for s in services:
                lines.append(
                    f'   - {_date(s.tanggal_service)} | {_display(s, "jenis_service")} | {s.bengkel or "-"} | '
                    f'{_rupiah_text(s.total_biaya)} | Kondisi: {_display(s, "kondisi_sebelum")} -> {_display(s, "kondisi_sesudah")}'
                )
        else:
            lines.append('   Riwayat service: -')
        lines.append('')

    return _pdf_response(lines, f'laporan_kendaraan_sip_service_{_today_label()}.pdf', 'Laporan Kendaraan')


# -----------------------------------------------------------------------------
# Export Laporan Rumah Dinas
# -----------------------------------------------------------------------------
def _get_latest_sip_rumah(rumah):
    return rumah.sip_rumah.all().order_by('-tanggal_sip').first()


@login_required
@user_passes_test(can_manage_master, login_url='login')
def export_rumah_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Rumah Dinas dan SIP'

    ws.append([
        'Kode Rumah', 'Nama Rumah', 'Jenis', 'Alamat', 'Provinsi', 'Kab/Kota',
        'Luas Tanah', 'Luas Bangunan', 'Kondisi', 'Status Pemanfaatan', 'Nilai Perolehan',
        'Nomor SIP Terakhir', 'Tanggal SIP', 'Pemakai', 'NIP Pemakai', 'Jabatan Pemakai',
        'Unit Kerja Pemakai', 'Periode SIP', 'Status SIP', 'Jumlah Anggota Keluarga',
        'Pejabat Penandatangan', 'Catatan SIP', 'Riwayat Pengguna Sebelumnya'
    ])

    rumah_qs = RumahDinas.objects.prefetch_related('sip_rumah__pegawai__unit_kerja').order_by('kode_rumah')
    for r in rumah_qs:
        sip = _get_latest_sip_rumah(r)
        pegawai = sip.pegawai if sip else None
        previous_sips = _get_previous_sips(
            list(r.sip_rumah.all().order_by('-tanggal_mulai', '-tanggal_sip')),
            sip
        )
        ws.append([
            r.kode_rumah, r.nama_rumah, _text(r.jenis_rumah), _text(r.alamat), _text(r.provinsi),
            _text(r.kabupaten_kota), _text(r.luas_tanah), _text(r.luas_bangunan), _display(r, 'kondisi'),
            _display(r, 'status_pemanfaatan'), _money(r.nilai_perolehan),
            _text(sip.nomor_sip if sip else None), _date(sip.tanggal_sip if sip else None),
            _text(pegawai.nama if pegawai else None), _text(pegawai.nip if pegawai else None),
            _text(pegawai.jabatan if pegawai else None), _text(pegawai.unit_kerja if pegawai else None),
            f"{_date(sip.tanggal_mulai)} s.d. {_date(sip.tanggal_akhir)}" if sip else '-',
            _display(sip, 'status') if sip else '-', _text(sip.jumlah_anggota_keluarga if sip else None),
            _text(sip.pejabat_penandatangan if sip else None), _text(sip.catatan if sip else None),
            _join_previous_users(previous_sips),
        ])
    _style_sheet(ws)

    ws_sip = wb.create_sheet('Detail SIP Rumah Dinas')
    ws_sip.append([
        'Nomor SIP', 'Tanggal SIP', 'Kode Rumah', 'Nama Rumah', 'Alamat', 'Kondisi Rumah',
        'Pemakai', 'NIP Pemakai', 'Jabatan', 'Unit Kerja', 'Tanggal Mulai', 'Tanggal Akhir',
        'Status SIP', 'Jumlah Anggota Keluarga', 'Pejabat Penandatangan', 'Dasar Penerbitan', 'Catatan'
    ])
    for sip in SIPRumahDinas.objects.select_related('rumah_dinas', 'pegawai__unit_kerja').order_by('-tanggal_sip'):
        ws_sip.append([
            sip.nomor_sip, _date(sip.tanggal_sip), sip.rumah_dinas.kode_rumah,
            sip.rumah_dinas.nama_rumah, sip.rumah_dinas.alamat, _display(sip.rumah_dinas, 'kondisi'),
            sip.pegawai.nama, sip.pegawai.nip, _text(sip.pegawai.jabatan), _text(sip.pegawai.unit_kerja),
            _date(sip.tanggal_mulai), _date(sip.tanggal_akhir), _display(sip, 'status'),
            sip.jumlah_anggota_keluarga, _text(sip.pejabat_penandatangan),
            _text(sip.dasar_penerbitan), _text(sip.catatan),
        ])
    _style_sheet(ws_sip)

    ws_pengguna = wb.create_sheet('Riwayat Pengguna Rumah')
    ws_pengguna.append([
        'Kode Rumah', 'Nama Rumah', 'Alamat', 'Nomor SIP', 'Tanggal SIP', 'Nama Pegawai',
        'NIP', 'Jabatan', 'Unit Kerja', 'Tanggal Mulai', 'Tanggal Akhir', 'Status SIP',
        'Jumlah Anggota Keluarga', 'Pejabat Penandatangan', 'Catatan'
    ])
    for sip in SIPRumahDinas.objects.select_related(
        'rumah_dinas', 'pegawai__unit_kerja'
    ).order_by('rumah_dinas__kode_rumah', '-tanggal_mulai', '-tanggal_sip'):
        ws_pengguna.append([
            sip.rumah_dinas.kode_rumah, sip.rumah_dinas.nama_rumah, sip.rumah_dinas.alamat,
            sip.nomor_sip, _date(sip.tanggal_sip), sip.pegawai.nama, sip.pegawai.nip,
            _text(sip.pegawai.jabatan), _text(sip.pegawai.unit_kerja),
            _date(sip.tanggal_mulai), _date(sip.tanggal_akhir), _display(sip, 'status'),
            sip.jumlah_anggota_keluarga, _text(sip.pejabat_penandatangan), _text(sip.catatan),
        ])
    _style_sheet(ws_pengguna)

    return _excel_response(wb, f'laporan_rumah_dinas_sip_pemakai_{_today_label()}.xlsx')


@login_required
@user_passes_test(can_manage_master, login_url='login')
def export_rumah_pdf(request):
    lines = []
    lines.append('LAPORAN RUMAH DINAS, SIP, KONDISI, DAN PEMAKAI')
    lines.append(f'Dicetak: {timezone.localtime().strftime("%d-%m-%Y %H:%M") }')
    lines.append('')

    rumah_qs = RumahDinas.objects.prefetch_related('sip_rumah__pegawai__unit_kerja').order_by('kode_rumah')
    for no, r in enumerate(rumah_qs, start=1):
        sip = _get_latest_sip_rumah(r)
        lines.append(f'{no}. {r.kode_rumah} - {r.nama_rumah}')
        lines.append(f'   Jenis: {_text(r.jenis_rumah)} | Kondisi: {_display(r, "kondisi")} | Status: {_display(r, "status_pemanfaatan")}')
        lines.append(f'   Alamat: {_text(r.alamat)}')
        lines.append(f'   Luas tanah/bangunan: {_text(r.luas_tanah)} / {_text(r.luas_bangunan)} | Nilai: {_rupiah_text(r.nilai_perolehan)}')
        if sip:
            pegawai = sip.pegawai
            lines.append(f'   SIP terakhir: {sip.nomor_sip} ({_display(sip, "status")}), tanggal {_date(sip.tanggal_sip)}')
            lines.append(f'   Pemakai: {pegawai.nama} / {pegawai.nip} | Jabatan: {_text(pegawai.jabatan)}')
            lines.append(f'   Unit kerja: {_text(pegawai.unit_kerja)} | Periode: {_date(sip.tanggal_mulai)} s.d. {_date(sip.tanggal_akhir)}')
            lines.append(f'   Jumlah anggota keluarga: {sip.jumlah_anggota_keluarga} | Catatan: {_text(sip.catatan)}')
        else:
            lines.append('   SIP/Pemakai: -')

        previous_sips = _get_previous_sips(
            list(r.sip_rumah.all().order_by('-tanggal_mulai', '-tanggal_sip')),
            sip
        )
        if previous_sips:
            lines.append('   Riwayat pengguna sebelumnya:')
            for prev_sip in previous_sips[:10]:
                lines.append(f'   - {_format_sip_history(prev_sip)}')
        else:
            lines.append('   Riwayat pengguna sebelumnya: -')
        lines.append('')

    return _pdf_response(lines, f'laporan_rumah_dinas_sip_pemakai_{_today_label()}.pdf', 'Laporan Rumah Dinas')

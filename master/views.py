from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse_lazy
from django.db.models.deletion import ProtectedError
from django.views.generic import CreateView, UpdateView, DetailView, DeleteView, FormView
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required

from core.roles import BMNRequiredMixin, VehicleViewRequiredMixin, can_manage_master, is_pengelola_bmn, is_admin_system
from core.forms import ImportFileForm
from core.import_utils import read_tabular_upload, pick, to_int, to_decimal, to_date, normalize_choice
from core.constants import KONDISI_ASET, STATUS_PEMANFAATAN_KENDARAAN, STATUS_PEMANFAATAN_RUMAH, JENIS_KENDARAAN_CHOICES, JENIS_UNIT_KERJA_CHOICES
from core.listing import SearchListMixin
from core.detail import GenericDetailMixin
from core.access import UnitScopedQuerysetMixin, UnitScopedFormMixin, BiroUmumOnlyMixin, scope_queryset_by_user, is_biro_umum_user, get_user_unit_kerja, require_user_unit_or_all, is_global_bmn_scope_user, get_accessible_unit_ids_for_user
from .models import (
    UnitKerja,
    Pegawai,
    Kendaraan,
    RumahDinas,
    FotoKendaraan,
    FotoRumahDinas,
)
from .forms import UnitKerjaForm, PegawaiForm, KendaraanForm, RumahDinasForm


class SafeDeleteMixin:
    """DeleteView helper untuk tombol hapus dengan pesan dan proteksi relasi."""
    template_name = 'includes/confirm_delete.html'
    success_message = 'Data berhasil dihapus.'
    protected_message = 'Data tidak dapat dihapus karena masih digunakan oleh data lain.'

    def form_valid(self, form):
        try:
            messages.success(self.request, self.success_message)
            return super().form_valid(form)
        except ProtectedError:
            messages.error(self.request, self.protected_message)
            return redirect(self.get_success_url())


class UnitKerjaListView(BMNRequiredMixin, UnitScopedQuerysetMixin, SearchListMixin):
    scope_type = 'unit'
    model = UnitKerja
    template_name = 'master/unitkerja_list.html'
    search_fields = [
        ('nama_unit', 'Nama Unit Kerja'),
        ('keterangan', 'Keterangan'),
    ]


class UnitKerjaCreateView(BMNRequiredMixin, BiroUmumOnlyMixin, UnitScopedFormMixin, CreateView):
    model = UnitKerja
    form_class = UnitKerjaForm
    template_name = 'master/form.html'
    success_url = reverse_lazy('master:unitkerja_list')


class UnitKerjaUpdateView(BMNRequiredMixin, UnitScopedQuerysetMixin, UnitScopedFormMixin, UpdateView):
    scope_type = 'unit'
    model = UnitKerja
    form_class = UnitKerjaForm
    template_name = 'master/form.html'
    success_url = reverse_lazy('master:unitkerja_list')


class UnitKerjaDetailView(BMNRequiredMixin, UnitScopedQuerysetMixin, GenericDetailMixin, DetailView):
    scope_type = 'unit'
    model = UnitKerja
    detail_title = 'Detail Unit Kerja'
    back_url_name = 'master:unitkerja_list'
    edit_url_name = 'master:unitkerja_update'
    delete_url_name = 'master:unitkerja_delete'


class UnitKerjaDeleteView(BMNRequiredMixin, BiroUmumOnlyMixin, UnitScopedQuerysetMixin, SafeDeleteMixin, DeleteView):
    scope_type = 'unit'
    model = UnitKerja
    success_url = reverse_lazy('master:unitkerja_list')
    success_message = 'Unit kerja berhasil dihapus.'
    protected_message = 'Unit kerja tidak dapat dihapus karena masih digunakan oleh pegawai/aset/transaksi.'


class PegawaiListView(BMNRequiredMixin, UnitScopedQuerysetMixin, SearchListMixin):
    scope_type = 'pegawai'
    model = Pegawai
    template_name = 'master/pegawai_list.html'
    select_related = ['unit_kerja']
    search_fields = [
        ('nip', 'NIP'),
        ('nik', 'NIK'),
        ('nama', 'Nama Pegawai'),
        ('jabatan', 'Jabatan'),
        ('pangkat', 'Pangkat'),
        ('golongan', 'Golongan'),
        ('unit_kerja__nama_unit', 'Unit Kerja'),
        ('no_hp', 'No HP'),
        ('email', 'Email'),
        ('status_pegawai', 'Status Pegawai'),
    ]


class PegawaiCreateView(BMNRequiredMixin, UnitScopedFormMixin, CreateView):
    model = Pegawai
    form_class = PegawaiForm
    template_name = 'master/pegawai_form.html'
    success_url = reverse_lazy('master:pegawai_list')


class PegawaiUpdateView(BMNRequiredMixin, UnitScopedQuerysetMixin, UnitScopedFormMixin, UpdateView):
    scope_type = 'pegawai'
    model = Pegawai
    form_class = PegawaiForm
    template_name = 'master/pegawai_form.html'
    success_url = reverse_lazy('master:pegawai_list')


class PegawaiDetailView(BMNRequiredMixin, UnitScopedQuerysetMixin, GenericDetailMixin, DetailView):
    scope_type = 'pegawai'
    model = Pegawai
    detail_title = 'Detail Pegawai'
    back_url_name = 'master:pegawai_list'
    edit_url_name = 'master:pegawai_update'
    delete_url_name = 'master:pegawai_delete'


class PegawaiDeleteView(BMNRequiredMixin, UnitScopedQuerysetMixin, SafeDeleteMixin, DeleteView):
    scope_type = 'pegawai'
    model = Pegawai
    success_url = reverse_lazy('master:pegawai_list')
    success_message = 'Pegawai berhasil dihapus.'
    protected_message = 'Pegawai tidak dapat dihapus karena masih digunakan pada SIP/pengguna aset/transaksi.'


@login_required
@require_POST
def pegawai_foto_delete(request, pk):
    if not can_manage_master(request.user):
        messages.error(request, 'Anda tidak memiliki hak akses untuk menghapus foto pegawai.')
        return redirect('dashboard')

    pegawai = get_object_or_404(scope_queryset_by_user(Pegawai.objects.all(), request.user, 'pegawai'), pk=pk)

    if pegawai.foto:
        pegawai.foto.delete(save=False)
        pegawai.foto = None
        pegawai.save(update_fields=['foto', 'updated_at'])
        messages.success(request, 'Foto pegawai berhasil dihapus.')
    else:
        messages.info(request, 'Pegawai ini belum memiliki foto.')

    return redirect('master:pegawai_update', pk=pegawai.pk)


class KendaraanListView(VehicleViewRequiredMixin, UnitScopedQuerysetMixin, SearchListMixin):
    scope_type = 'kendaraan'
    model = Kendaraan
    template_name = 'master/kendaraan_list.html'
    select_related = ['unit_kerja']
    search_fields = [
        ('kode_kendaraan', 'Kode Kendaraan'),
        ('nomor_polisi', 'Nomor Polisi'),
        ('merek', 'Merek'),
        ('tipe', 'Tipe'),
        ('jenis_kendaraan', 'Jenis Kendaraan'),
        ('warna', 'Warna'),
        ('nomor_rangka', 'Nomor Rangka'),
        ('nomor_mesin', 'Nomor Mesin'),
        ('nomor_bpkb', 'Nomor BPKB'),
        ('nomor_stnk', 'Nomor STNK'),
        ('nup', 'NUP'),
        ('kode_barang', 'Kode Barang'),
        ('unit_kerja__nama_unit', 'Unit Kerja'),
        ('kondisi', 'Kondisi'),
        ('status_pemanfaatan', 'Status Pemanfaatan'),
    ]


class KendaraanPhotoMixin:
    template_name = 'master/kendaraan_form.html'
    success_url = reverse_lazy('master:kendaraan_list')

    def save_uploaded_photos(self, kendaraan):
        for foto in self.request.FILES.getlist('foto_kendaraan'):
            FotoKendaraan.objects.create(
                kendaraan=kendaraan,
                foto=foto,
                diupload_oleh=self.request.user if self.request.user.is_authenticated else None
            )

    def form_valid(self, form):
        response = super().form_valid(form)
        self.save_uploaded_photos(self.object)
        messages.success(self.request, 'Data kendaraan berhasil disimpan.')
        return response

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        if getattr(self, 'object', None):
            ctx['foto_kendaraan_list'] = self.object.galeri_foto.all()
        else:
            ctx['foto_kendaraan_list'] = []

        return ctx


class KendaraanCreateView(BMNRequiredMixin, UnitScopedFormMixin, KendaraanPhotoMixin, CreateView):
    model = Kendaraan
    form_class = KendaraanForm


class KendaraanUpdateView(BMNRequiredMixin, UnitScopedQuerysetMixin, UnitScopedFormMixin, KendaraanPhotoMixin, UpdateView):
    scope_type = 'kendaraan'
    model = Kendaraan
    form_class = KendaraanForm


class KendaraanDetailView(VehicleViewRequiredMixin, UnitScopedQuerysetMixin, GenericDetailMixin, DetailView):
    scope_type = 'kendaraan'
    model = Kendaraan
    detail_title = 'Detail Kendaraan'
    back_url_name = 'master:kendaraan_list'
    edit_url_name = 'master:kendaraan_update'
    delete_url_name = 'master:kendaraan_delete'
    exclude_fields = ['id', 'pengguna']


class KendaraanDeleteView(BMNRequiredMixin, UnitScopedQuerysetMixin, SafeDeleteMixin, DeleteView):
    scope_type = 'kendaraan'
    model = Kendaraan
    success_url = reverse_lazy('master:kendaraan_list')
    success_message = 'Kendaraan berhasil dihapus.'
    protected_message = 'Kendaraan tidak dapat dihapus karena masih digunakan pada SIP/service/riwayat kondisi.'


@login_required
@require_POST
def kendaraan_foto_delete(request, pk):
    if not can_manage_master(request.user):
        messages.error(request, 'Anda tidak memiliki hak akses untuk menghapus foto kendaraan.')
        return redirect('dashboard')

    foto = get_object_or_404(FotoKendaraan.objects.filter(kendaraan__in=scope_queryset_by_user(Kendaraan.objects.all(), request.user, 'kendaraan')), pk=pk)
    kendaraan_id = foto.kendaraan_id
    foto.foto.delete(save=False)
    foto.delete()

    messages.success(request, 'Foto kendaraan berhasil dihapus.')
    return redirect('master:kendaraan_update', pk=kendaraan_id)


class RumahDinasListView(BMNRequiredMixin, UnitScopedQuerysetMixin, SearchListMixin):
    scope_type = 'rumah'
    model = RumahDinas
    template_name = 'master/rumah_list.html'
    search_fields = [
        ('kode_rumah', 'Kode Rumah'),
        ('nama_rumah', 'Nama Rumah'),
        ('jenis_rumah', 'Jenis Rumah'),
        ('alamat', 'Alamat'),
        ('provinsi', 'Provinsi'),
        ('kabupaten_kota', 'Kabupaten/Kota'),
        ('kecamatan', 'Kecamatan'),
        ('kelurahan', 'Kelurahan'),
        ('nup', 'NUP'),
        ('kode_barang', 'Kode Barang'),
        ('nomor_sertifikat', 'Nomor Sertifikat'),
        ('status_tanah', 'Status Tanah'),
        ('unit_kerja__nama_unit', 'Unit Kerja'),
        ('kondisi', 'Kondisi'),
        ('status_pemanfaatan', 'Status Pemanfaatan'),
    ]


class RumahDinasPhotoMixin:
    template_name = 'master/rumah_form.html'
    success_url = reverse_lazy('master:rumah_list')

    def save_uploaded_photos(self, rumah_dinas):
        for foto in self.request.FILES.getlist('foto_rumah_dinas'):
            FotoRumahDinas.objects.create(
                rumah_dinas=rumah_dinas,
                foto=foto,
                diupload_oleh=self.request.user if self.request.user.is_authenticated else None
            )

    def form_valid(self, form):
        response = super().form_valid(form)
        self.save_uploaded_photos(self.object)
        messages.success(self.request, 'Data rumah negara berhasil disimpan.')
        return response

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        if getattr(self, 'object', None):
            ctx['foto_rumah_list'] = self.object.galeri_foto.all()
        else:
            ctx['foto_rumah_list'] = []

        return ctx


class RumahDinasCreateView(BMNRequiredMixin, UnitScopedFormMixin, RumahDinasPhotoMixin, CreateView):
    model = RumahDinas
    form_class = RumahDinasForm


class RumahDinasUpdateView(BMNRequiredMixin, UnitScopedQuerysetMixin, UnitScopedFormMixin, RumahDinasPhotoMixin, UpdateView):
    scope_type = 'rumah'
    model = RumahDinas
    form_class = RumahDinasForm


class RumahDinasDetailView(BMNRequiredMixin, UnitScopedQuerysetMixin, GenericDetailMixin, DetailView):
    scope_type = 'rumah'
    model = RumahDinas
    detail_title = 'Detail Rumah Negara'
    back_url_name = 'master:rumah_list'
    edit_url_name = 'master:rumah_update'
    delete_url_name = 'master:rumah_delete'


class RumahDinasDeleteView(BMNRequiredMixin, UnitScopedQuerysetMixin, SafeDeleteMixin, DeleteView):
    scope_type = 'rumah'
    model = RumahDinas
    success_url = reverse_lazy('master:rumah_list')
    success_message = 'Rumah negara berhasil dihapus.'
    protected_message = 'Rumah negara tidak dapat dihapus karena masih digunakan pada SIP/transaksi terkait.'


@login_required
@require_POST
def rumah_foto_delete(request, pk):
    if not can_manage_master(request.user):
        messages.error(request, 'Anda tidak memiliki hak akses untuk menghapus foto rumah negara.')
        return redirect('dashboard')

    foto = get_object_or_404(FotoRumahDinas.objects.filter(rumah_dinas__in=scope_queryset_by_user(RumahDinas.objects.all(), request.user, 'rumah')), pk=pk)
    rumah_id = foto.rumah_dinas_id
    foto.foto.delete(save=False)
    foto.delete()

    messages.success(request, 'Foto rumah negara berhasil dihapus.')
    return redirect('master:rumah_update', pk=rumah_id)


# =========================
# EXPORT MASTER DATA PDF/EXCEL/CSV
# =========================

def _stringify(value):
    if value is None:
        return ''
    if hasattr(value, 'strftime'):
        try:
            return value.strftime('%d/%m/%Y')
        except Exception:
            pass
    return str(value)


def _apply_search_filter(qs, request, search_fields):
    q = (request.GET.get('q') or '').strip()
    selected_field = (request.GET.get('search_field') or 'ALL').strip()
    if q and search_fields:
        available_fields = [field for field, _label in search_fields]
        fields_to_search = available_fields if selected_field == 'ALL' or selected_field not in available_fields else [selected_field]
        query = Q()
        for field in fields_to_search:
            query |= Q(**{f'{field}__icontains': q})
        qs = qs.filter(query)
    return qs


def _get_export_queryset(request, model, scope_type, search_fields, select_related=None):
    qs = model.objects.all()
    if select_related:
        qs = qs.select_related(*select_related)
    qs = scope_queryset_by_user(qs, request.user, scope_type)
    return _apply_search_filter(qs, request, search_fields)


def _export_csv_response(filename, headers, rows):
    import csv
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow([_stringify(v) for v in row])
    return response


def _export_excel_response(filename, title, headers, rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        return HttpResponse(f'openpyxl belum tersedia: {exc}', status=500, content_type='text/plain')

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(headers)
    header_fill = PatternFill('solid', fgColor='1D4ED8')
    header_font = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    for row in rows:
        ws.append([_stringify(v) for v in row])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    for idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for cell in ws[get_column_letter(idx)]:
            max_len = max(max_len, len(str(cell.value or '')))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 38)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _export_pdf_response(filename, title, headers, rows, landscape_mode=True):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except Exception as exc:
        return HttpResponse(f'reportlab belum tersedia: {exc}', status=500, content_type='text/plain')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    page_size = landscape(A4) if landscape_mode else A4
    doc = SimpleDocTemplate(response, pagesize=page_size, leftMargin=0.8*cm, rightMargin=0.8*cm, topMargin=0.8*cm, bottomMargin=0.8*cm)
    styles = getSampleStyleSheet()
    elements = [Paragraph(f'<b>{title}</b>', styles['Title']), Spacer(1, 0.25*cm)]
    max_rows = 500
    table_rows = [headers]
    for row in rows[:max_rows]:
        table_rows.append([Paragraph(_stringify(v), styles['BodyText']) for v in row])
    table = Table(table_rows, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1D4ED8')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5E1')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
    ]))
    elements.append(table)
    if len(rows) > max_rows:
        elements.append(Spacer(1, 0.2*cm))
        elements.append(Paragraph(f'Catatan: PDF dibatasi {max_rows} baris pertama. Gunakan Excel/CSV untuk data lengkap.', styles['Normal']))
    doc.build(elements)
    return response


def _unit_rows(qs):
    return [[i, o.nama_unit, getattr(o, 'jenis_unit', ''), getattr(o, 'nama_jabatan_penerbit_sip_kendaraan', '') or '', o.keterangan or ''] for i, o in enumerate(qs, start=1)]


def _pegawai_rows(qs):
    return [[o.nip, o.nama, o.jabatan or '', o.pangkat or '', o.golongan or '', o.unit_kerja.nama_unit if o.unit_kerja else '', o.no_hp or '', o.email or '', o.status_pegawai or ''] for o in qs]


def _kendaraan_rows(qs):
    return [[o.nomor_polisi, o.merek or '', o.tipe or '', o.jenis_kendaraan or '', o.tahun_perolehan or '', o.unit_kerja.nama_unit if o.unit_kerja else '', o.nup or '', o.kode_barang or '', o.nilai_perolehan or 0, o.kondisi or '', o.status_pemanfaatan or ''] for o in qs]


def _rumah_rows(qs):
    return [[o.kode_rumah, o.nama_rumah or '', o.jenis_rumah or '', o.unit_kerja.nama_unit if o.unit_kerja else '', o.alamat or '', o.nup or '', o.kode_barang or '', o.nilai_perolehan or 0, o.kondisi or '', o.status_pemanfaatan or ''] for o in qs]


@login_required
def export_unitkerja(request, fmt):
    if not can_manage_master(request.user):
        messages.error(request, 'Anda tidak memiliki hak akses export Unit Kerja.')
        return redirect('dashboard')
    fields = UnitKerjaListView.search_fields
    qs = _get_export_queryset(request, UnitKerja, 'unit', fields).order_by('nama_unit')
    headers = ['No', 'Nama Unit Kerja', 'Jenis Unit', 'Jabatan Penerbit SIP Kendaraan', 'Keterangan']
    rows = _unit_rows(qs)
    if fmt == 'pdf':
        return _export_pdf_response('master_unit_kerja.pdf', 'Master Unit Kerja', headers, rows)
    if fmt == 'csv':
        return _export_csv_response('master_unit_kerja.csv', headers, rows)
    return _export_excel_response('master_unit_kerja.xlsx', 'Unit Kerja', headers, rows)


@login_required
def export_pegawai(request, fmt):
    if not can_manage_master(request.user):
        messages.error(request, 'Anda tidak memiliki hak akses export Pegawai.')
        return redirect('dashboard')
    fields = PegawaiListView.search_fields
    qs = _get_export_queryset(request, Pegawai, 'pegawai', fields, ['unit_kerja']).order_by('unit_kerja__nama_unit', 'nama')
    headers = ['NIP', 'Nama', 'Jabatan', 'Pangkat', 'Golongan', 'Unit Kerja', 'No HP', 'Email', 'Status']
    rows = _pegawai_rows(qs)
    if fmt == 'pdf':
        return _export_pdf_response('master_pegawai.pdf', 'Master Pegawai', headers, rows)
    if fmt == 'csv':
        return _export_csv_response('master_pegawai.csv', headers, rows)
    return _export_excel_response('master_pegawai.xlsx', 'Pegawai', headers, rows)


@login_required
def export_kendaraan(request, fmt):
    if not can_manage_master(request.user):
        messages.error(request, 'Anda tidak memiliki hak akses export Kendaraan.')
        return redirect('dashboard')
    fields = KendaraanListView.search_fields
    qs = _get_export_queryset(request, Kendaraan, 'kendaraan', fields, ['unit_kerja']).order_by('unit_kerja__nama_unit', 'nomor_polisi')
    headers = ['No Polisi', 'Merek', 'Tipe', 'Jenis', 'Tahun', 'Unit Kerja', 'NUP', 'Kode Barang', 'Nilai Perolehan', 'Kondisi', 'Status Pemanfaatan']
    rows = _kendaraan_rows(qs)
    if fmt == 'pdf':
        return _export_pdf_response('master_kendaraan.pdf', 'Master Kendaraan', headers, rows)
    if fmt == 'csv':
        return _export_csv_response('master_kendaraan.csv', headers, rows)
    return _export_excel_response('master_kendaraan.xlsx', 'Kendaraan', headers, rows)


@login_required
def export_rumah(request, fmt):
    if not can_manage_master(request.user):
        messages.error(request, 'Anda tidak memiliki hak akses export Rumah Negara.')
        return redirect('dashboard')
    fields = RumahDinasListView.search_fields
    qs = _get_export_queryset(request, RumahDinas, 'rumah', fields, ['unit_kerja']).order_by('unit_kerja__nama_unit', 'kode_rumah')
    headers = ['Kode', 'Nama Rumah', 'Jenis', 'Unit Kerja', 'Alamat', 'NUP', 'Kode Barang', 'Nilai Perolehan', 'Kondisi', 'Status Pemanfaatan']
    rows = _rumah_rows(qs)
    if fmt == 'pdf':
        return _export_pdf_response('master_rumah_negara.pdf', 'Master Rumah Negara', headers, rows)
    if fmt == 'csv':
        return _export_csv_response('master_rumah_negara.csv', headers, rows)
    return _export_excel_response('master_rumah_negara.xlsx', 'Rumah Negara', headers, rows)


# =========================
# IMPORT EXCEL/CSV
# =========================

def _is_ajax(request):
    return request.headers.get('x-requested-with') == 'XMLHttpRequest'


def _json_or_redirect(request, ok, message, redirect_url, status=200):
    if _is_ajax(request):
        return JsonResponse({'ok': ok, 'message': message, 'redirect_url': str(redirect_url)}, status=status)
    if ok:
        messages.success(request, message)
    else:
        messages.error(request, message)
    return redirect(redirect_url)


def _get_or_create_unit(name, user):
    """Unit kerja untuk import master data.

    Admin System/Biro Umum boleh menentukan/membuat unit kerja dari file import.
    Pengelola BMN Sekretariat kantor pusat boleh mengimpor ke unit yang masih
    berada dalam scope Eselon I-nya. BMN Sentra/Balai tetap dipaksa ke unitnya sendiri.
    """
    if is_global_bmn_scope_user(user):
        if not name:
            return None
        return UnitKerja.objects.get_or_create(nama_unit=str(name).strip())[0]

    own_unit = get_user_unit_kerja(user)
    if not own_unit:
        return None

    if name:
        nama = str(name).strip()
        allowed_ids = get_accessible_unit_ids_for_user(user) or [own_unit.pk]
        existing = UnitKerja.objects.filter(pk__in=allowed_ids, nama_unit__iexact=nama).first()
        if existing:
            return existing

    return own_unit




# =============================================================
# Template Excel Import Master Data
# =============================================================
def _excel_import_template_response(filename, title, headers, sample_row, notes=None):
    """Generate template Excel sederhana untuk import data master."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        return HttpResponse(f'openpyxl belum tersedia: {exc}', status=500, content_type='text/plain')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Template Import'

    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14, color='0F172A')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))

    ws.append([])
    ws.append(headers)
    ws.append(sample_row)

    header_fill = PatternFill('solid', fgColor='1D4ED8')
    header_font = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    for cell in ws[4]:
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

    for idx, header in enumerate(headers, start=1):
        width = min(max(len(str(header)) + 4, 14), 32)
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:{get_column_letter(len(headers))}4'

    info = wb.create_sheet('Petunjuk')
    info['A1'] = 'Petunjuk Pengisian Template Import'
    info['A1'].font = Font(bold=True, size=14, color='0F172A')
    info['A3'] = '1. Jangan mengubah nama header/kolom pada baris ke-3.'
    info['A4'] = '2. Isi data mulai baris ke-4. Baris contoh boleh dihapus.'
    info['A5'] = '3. Kolom unit_kerja harus sesuai nama unit kerja pada aplikasi.'
    info['A6'] = '4. Format tanggal disarankan YYYY-MM-DD, contoh 2026-06-15.'
    info['A7'] = '5. Untuk nilai rupiah gunakan angka saja tanpa Rp/titik/koma.'
    r = 9
    if notes:
        info[f'A{r}'] = 'Catatan khusus:'
        info[f'A{r}'].font = Font(bold=True)
        for note in notes:
            r += 1
            info[f'A{r}'] = f'- {note}'
    info.column_dimensions['A'].width = 110
    for row in info.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def template_import_unitkerja(request):
    headers = [
        'nama_unit',
        'jenis_unit',
        'nama_jabatan_penerbit_sip_kendaraan',
        'pejabat_penerbit_nip',
        'keterangan',
    ]
    sample = [
        'Biro Umum',
        'BIRO_UMUM',
        'Kepala Biro Umum',
        '197004281998031004',
        'Unit kerja Sekretariat Jenderal',
    ]
    notes = [
        'Kolom nama_unit wajib diisi dan menjadi kunci update data.',
        'jenis_unit dapat diisi: BIRO_UMUM, DITJEN, ITJEN, BADAN, PUSAT, SENTRA, BALAI, atau LAINNYA.',
        'pejabat_penerbit_nip opsional. Isi NIP pegawai yang sudah ada pada Master Pegawai jika ingin mengatur pejabat penerbit SIP Kendaraan.',
        'Jika pejabat_penerbit_nip kosong atau tidak ditemukan, data unit kerja tetap diimpor tanpa pejabat penerbit.',
    ]
    return _excel_import_template_response('template_import_unit_kerja.xlsx', 'Template Import Unit Kerja', headers, sample, notes)


def template_import_pegawai(request):
    headers = ['nip', 'nik', 'nama', 'jabatan', 'pangkat', 'golongan', 'unit_kerja', 'no_hp', 'email', 'alamat', 'status_pegawai']
    sample = ['198001012006041001', '3173010101800001', 'Contoh Pegawai', 'Analis Barang Milik Negara', 'Penata Muda Tk. I', 'III/b', 'Biro Umum', '081234567890', 'contoh@kemsos.go.id', 'Jakarta', 'Aktif']
    return _excel_import_template_response('template_import_pegawai.xlsx', 'Template Import Pegawai', headers, sample)


def template_import_kendaraan(request):
    headers = ['kode_kendaraan', 'nomor_polisi', 'merek', 'tipe', 'jenis_kendaraan', 'tahun_pembuatan', 'tahun_perolehan', 'warna', 'nomor_rangka', 'nomor_mesin', 'nomor_bpkb', 'nomor_stnk', 'masa_berlaku_stnk', 'jatuh_tempo_pajak', 'nup', 'kode_barang', 'nilai_perolehan', 'unit_kerja', 'kondisi', 'status_pemanfaatan', 'kilometer_terakhir']
    sample = ['KDR-001', 'B 1234 KMS', 'Toyota', 'Innova', 'Operasional', 2022, 2022, 'Hitam', 'MHFXW42G...', '2GD123...', 'BPKB001', 'STNK001', '2027-06-15', '2026-12-31', '001', '3.02.01.02.003', 350000000, 'Biro Umum', 'Baik', 'Digunakan', 25000]
    notes = ['Pengguna kendaraan tidak diisi pada master kendaraan, karena pengguna dicatat melalui SIP Kendaraan.']
    return _excel_import_template_response('template_import_kendaraan.xlsx', 'Template Import Aset Kendaraan', headers, sample, notes)


def template_import_rumah(request):
    headers = ['kode_rumah', 'nama_rumah', 'jenis_rumah', 'alamat', 'provinsi', 'kabupaten_kota', 'kecamatan', 'kelurahan', 'latitude', 'longitude', 'luas_tanah', 'luas_bangunan', 'jumlah_kamar_tidur', 'jumlah_kamar_mandi', 'daya_listrik', 'tahun_dibangun', 'tahun_perolehan', 'nup', 'kode_barang', 'nilai_perolehan', 'unit_kerja', 'nomor_sertifikat', 'status_tanah', 'kondisi', 'status_pemanfaatan']
    sample = ['RN-001', 'Rumah Negara Contoh', 'Rumah Negara Golongan II', 'Jl. Contoh No. 1', 'DKI Jakarta', 'Jakarta Pusat', 'Senen', 'Senen', -6.175392, 106.827153, 120, 80, 3, 2, '2200 VA', 2010, 2011, '001', '4.01.01.01.001', 750000000, 'Biro Umum', 'SHM-001', 'Sertifikat', 'Baik', 'Kosong']
    return _excel_import_template_response('template_import_rumah_negara.xlsx', 'Template Import Aset Rumah Negara', headers, sample)


class BaseImportView(BMNRequiredMixin, FormView):
    form_class = ImportFileForm
    template_name = 'includes/import_form.html'
    title = 'Impor Data'
    description = ''
    back_url_name = None
    template_url_name = None

    def get_success_url(self):
        return reverse_lazy(self.back_url_name)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update({'title': self.title, 'description': self.description, 'back_url': self.get_success_url(), 'template_url': reverse_lazy(self.template_url_name) if self.template_url_name else None})
        return ctx

    def process_rows(self, rows):
        raise NotImplementedError

    def form_valid(self, form):
        try:
            rows = read_tabular_upload(form.cleaned_data['file'])
            created, updated, errors = self.process_rows(rows)
            msg = f'{self.title} selesai. Baru: {created}, Update: {updated}, Gagal: {errors}.'
            return _json_or_redirect(self.request, True, msg, self.get_success_url())
        except Exception as exc:
            return _json_or_redirect(self.request, False, str(exc), self.get_success_url(), status=400)

    def form_invalid(self, form):
        errors = []
        for field, field_errors in form.errors.items():
            for error in field_errors:
                errors.append(f'{field}: {error}')
        message = 'Form tidak valid. ' + '; '.join(errors)
        return _json_or_redirect(self.request, False, message, self.get_success_url(), status=400)


class UnitKerjaImportView(BiroUmumOnlyMixin, BaseImportView):
    title = 'Impor Unit Kerja'
    template_url_name = 'master:unitkerja_template_import'
    description = 'Kolom yang didukung: nama_unit, jenis_unit, nama_jabatan_penerbit_sip_kendaraan, pejabat_penerbit_nip, keterangan. File dapat berupa Excel (.xlsx/.xlsm) atau CSV.'
    back_url_name = 'master:unitkerja_list'

    def process_rows(self, rows):
        created = updated = errors = 0
        for row in rows:
            nama_unit = pick(row, 'nama_unit', 'unit_kerja', 'nama', 'satker')
            if not nama_unit:
                errors += 1
                continue

            jenis_unit = normalize_choice(
                pick(row, 'jenis_unit', 'jenis', 'tipe_unit', default='LAINNYA'),
                JENIS_UNIT_KERJA_CHOICES,
                'LAINNYA'
            )

            pejabat = None
            pejabat_nip = pick(row, 'pejabat_penerbit_nip', 'nip_pejabat_penerbit', 'nip_pejabat')
            pejabat_nama = pick(row, 'pejabat_penerbit_nama', 'nama_pejabat_penerbit', 'nama_pejabat')
            if pejabat_nip:
                pejabat = Pegawai.objects.filter(nip=str(pejabat_nip).strip()).first()
            if pejabat is None and pejabat_nama:
                pejabat = Pegawai.objects.filter(nama__iexact=str(pejabat_nama).strip()).first()

            defaults = {
                'jenis_unit': jenis_unit,
                'nama_jabatan_penerbit_sip_kendaraan': pick(
                    row,
                    'nama_jabatan_penerbit_sip_kendaraan',
                    'jabatan_penerbit',
                    'jabatan_pejabat_penerbit',
                    default=''
                ) or None,
                'keterangan': pick(row, 'keterangan', 'catatan'),
            }
            if pejabat is not None:
                defaults['pejabat_penerbit_sip_kendaraan'] = pejabat

            obj, is_created = UnitKerja.objects.update_or_create(
                nama_unit=str(nama_unit).strip(),
                defaults=defaults
            )
            created += 1 if is_created else 0
            updated += 0 if is_created else 1
        return created, updated, errors


class PegawaiImportView(BaseImportView):
    title = 'Impor Pegawai'
    template_url_name = 'master:pegawai_template_import'
    description = 'Kolom yang didukung: nip, nik, nama, jabatan, pangkat, golongan, unit_kerja, no_hp, email, alamat, status_pegawai.'
    back_url_name = 'master:pegawai_list'

    def process_rows(self, rows):
        created = updated = errors = 0
        for row in rows:
            nip = pick(row, 'nip')
            nama = pick(row, 'nama', 'nama_pegawai')
            if not nip or not nama:
                errors += 1; continue
            obj, is_created = Pegawai.objects.update_or_create(
                nip=str(nip),
                defaults={
                    'nik': pick(row, 'nik'), 'nama': nama, 'jabatan': pick(row, 'jabatan'),
                    'pangkat': pick(row, 'pangkat'), 'golongan': pick(row, 'golongan'),
                    'unit_kerja': _get_or_create_unit(pick(row, 'unit_kerja', 'satker'), self.request.user),
                    'no_hp': pick(row, 'no_hp', 'hp', 'telepon'), 'email': pick(row, 'email'),
                    'alamat': pick(row, 'alamat'), 'status_pegawai': pick(row, 'status_pegawai', 'status', default='Aktif'),
                })
            created += 1 if is_created else 0; updated += 0 if is_created else 1
        return created, updated, errors


class KendaraanImportView(BaseImportView):
    title = 'Impor Aset Kendaraan'
    template_url_name = 'master:kendaraan_template_import'
    description = 'Kolom yang didukung: kode_kendaraan, nomor_polisi, merek, tipe, jenis_kendaraan, tahun_pembuatan, tahun_perolehan, warna, nomor_rangka, nomor_mesin, nomor_bpkb, nomor_stnk, masa_berlaku_stnk, jatuh_tempo_pajak, nup, kode_barang, nilai_perolehan, unit_kerja, kondisi, status_pemanfaatan, kilometer_terakhir. Pengguna kendaraan dicatat melalui SIP Kendaraan, bukan master kendaraan.'
    back_url_name = 'master:kendaraan_list'

    def process_rows(self, rows):
        created = updated = errors = 0
        for row in rows:
            kode = pick(row, 'kode_kendaraan', 'kode', 'nup')
            nopol = pick(row, 'nomor_polisi', 'nopol', 'plat_nomor')
            merek = pick(row, 'merek', 'merk', default='-')
            if not kode or not nopol:
                errors += 1; continue
            obj, is_created = Kendaraan.objects.update_or_create(
                nomor_polisi=str(nopol).strip().upper(),
                defaults={
                    'kode_kendaraan': str(kode), 'merek': merek, 'tipe': pick(row, 'tipe'),
                    'jenis_kendaraan': normalize_choice(pick(row, 'jenis_kendaraan'), JENIS_KENDARAAN_CHOICES, None),
                    'tahun_pembuatan': to_int(pick(row, 'tahun_pembuatan')), 'tahun_perolehan': to_int(pick(row, 'tahun_perolehan')),
                    'warna': pick(row, 'warna'), 'nomor_rangka': pick(row, 'nomor_rangka'), 'nomor_mesin': pick(row, 'nomor_mesin'),
                    'nomor_bpkb': pick(row, 'nomor_bpkb'), 'nomor_stnk': pick(row, 'nomor_stnk'),
                    'masa_berlaku_stnk': to_date(pick(row, 'masa_berlaku_stnk')), 'jatuh_tempo_pajak': to_date(pick(row, 'jatuh_tempo_pajak')),
                    'nup': pick(row, 'nup'), 'kode_barang': pick(row, 'kode_barang'), 'nilai_perolehan': to_decimal(pick(row, 'nilai_perolehan', 'nilai'), 0),
                    'unit_kerja': _get_or_create_unit(pick(row, 'unit_kerja', 'satker'), self.request.user),
                    'kondisi': normalize_choice(pick(row, 'kondisi'), KONDISI_ASET, 'BAIK'),
                    'status_pemanfaatan': normalize_choice(pick(row, 'status_pemanfaatan', 'status'), STATUS_PEMANFAATAN_KENDARAAN, 'TERSEDIA'),
                    'kilometer_terakhir': to_int(pick(row, 'kilometer_terakhir', 'km'), 0) or 0,
                })
            created += 1 if is_created else 0; updated += 0 if is_created else 1
        return created, updated, errors


class RumahNegaraImportView(BaseImportView):
    title = 'Impor Aset Rumah Negara'
    template_url_name = 'master:rumah_template_import'
    description = 'Kolom yang didukung: kode_rumah, nama_rumah, jenis_rumah, alamat, provinsi, kabupaten_kota, kecamatan, kelurahan, latitude, longitude, luas_tanah, luas_bangunan, nup, kode_barang, nilai_perolehan, nomor_sertifikat, status_tanah, kondisi, status_pemanfaatan.'
    back_url_name = 'master:rumah_list'

    def process_rows(self, rows):
        created = updated = errors = 0
        for row in rows:
            kode = pick(row, 'kode_rumah', 'kode', 'nup')
            nama = pick(row, 'nama_rumah', 'nama', 'nama_aset', default=str(kode or ''))
            alamat = pick(row, 'alamat', default='-')
            if not kode:
                errors += 1; continue
            obj, is_created = RumahDinas.objects.update_or_create(
                kode_rumah=str(kode),
                defaults={
                    'nama_rumah': nama, 'jenis_rumah': pick(row, 'jenis_rumah'), 'alamat': alamat,
                    'provinsi': pick(row, 'provinsi'), 'kabupaten_kota': pick(row, 'kabupaten_kota','kabupaten','kota'),
                    'kecamatan': pick(row, 'kecamatan'), 'kelurahan': pick(row, 'kelurahan'),
                    'latitude': to_decimal(pick(row, 'latitude', 'lat'), None), 'longitude': to_decimal(pick(row, 'longitude','long','lng'), None),
                    'luas_tanah': to_decimal(pick(row, 'luas_tanah'), None), 'luas_bangunan': to_decimal(pick(row, 'luas_bangunan'), None),
                    'jumlah_kamar_tidur': to_int(pick(row, 'jumlah_kamar_tidur','kamar_tidur'), 0) or 0,
                    'jumlah_kamar_mandi': to_int(pick(row, 'jumlah_kamar_mandi','kamar_mandi'), 0) or 0,
                    'daya_listrik': pick(row, 'daya_listrik'), 'tahun_dibangun': to_int(pick(row, 'tahun_dibangun')),
                    'tahun_perolehan': to_int(pick(row, 'tahun_perolehan')), 'nup': pick(row, 'nup'), 'kode_barang': pick(row, 'kode_barang'),
                    'nilai_perolehan': to_decimal(pick(row, 'nilai_perolehan','nilai'), 0), 'unit_kerja': _get_or_create_unit(pick(row, 'unit_kerja', 'satker'), self.request.user),
                    'nomor_sertifikat': pick(row, 'nomor_sertifikat'),
                    'status_tanah': pick(row, 'status_tanah'), 'kondisi': normalize_choice(pick(row, 'kondisi'), KONDISI_ASET, 'BAIK'),
                    'status_pemanfaatan': normalize_choice(pick(row, 'status_pemanfaatan', 'status'), STATUS_PEMANFAATAN_RUMAH, 'KOSONG'),
                })
            created += 1 if is_created else 0; updated += 0 if is_created else 1
        return created, updated, errors
